from io import BytesIO

from openpyxl import load_workbook

from app.extensions import SessionLocal
from app.models import ProjectRiskSummaryRecord


def unwrap(response):
    assert response.status_code < 400, response.get_data(as_text=True)
    payload = response.get_json()
    assert payload["code"] == "SUCCESS"
    return payload["data"]


def create_project(client, project_code="ENST-TEST-RISK-EXCEL"):
    return unwrap(
        client.post(
            "/api/v1/projects",
            json={
                "projectName": "Risk Excel Project",
                "projectCode": project_code,
                "assessmentOrg": "Test Org",
                "assessmentTemplateId": "tpl-gb",
                "scoreModelId": "score-v1",
                "harmModelId": "harm-default",
                "riskMatrixId": "matrix-v1",
                "systemType": "MANAGEMENT_SYSTEM",
            },
        )
    )["projectId"]


def workbook_stream(workbook):
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def seed_risk_summary(client, project_code="ENST-TEST-RISK-EXCEL"):
    project_id = create_project(client, project_code=project_code)
    unwrap(client.post(f"/api/v1/projects/{project_id}/start"))
    items = unwrap(client.get(f"/api/v1/projects/{project_id}/evaluation/items?pageNo=1&pageSize=1"))
    item_id = items["list"][0]["id"]
    unwrap(
        client.put(
            f"/api/v1/projects/{project_id}/evaluation/items/{item_id}/record",
            json={"evaluationResult": "PARTIAL", "evaluationRecord": "现场测评记录"},
        )
    )
    unwrap(client.post(f"/api/v1/projects/{project_id}/risk-summary/refresh", json={}))
    source_page = unwrap(client.get(f"/api/v1/projects/{project_id}/risk-sources?pageNo=1&pageSize=1"))
    return project_id, source_page["list"][0]["id"]


def test_risk_source_excel_export_and_import_updates_editable_fields(client):
    project_id, risk_record_id = seed_risk_summary(client, "ENST-TEST-RISK-SOURCE-EXCEL")

    export_response = client.get(f"/api/v1/projects/{project_id}/risk-sources/export")
    assert export_response.status_code == 200
    workbook = load_workbook(BytesIO(export_response.data))
    worksheet = workbook.active

    assert [cell.value for cell in worksheet[1][:11]] == [
        "序号",
        "评估项ID",
        "评估子类",
        "检查要点",
        "评估结果",
        "风险描述",
        "风险源描述",
        "风险源类型",
        "风险类型",
        "涉及的数据及类型、级别",
        "涉及的数据处理活动",
    ]
    assert worksheet.cell(row=1, column=12).value == "riskSourceId"
    assert worksheet.column_dimensions["L"].hidden is True
    assert worksheet.cell(row=2, column=2).value == "AQGL001"
    assert worksheet.cell(row=2, column=3).value == "数据安全制度体系"
    assert worksheet.cell(row=2, column=8).value == "数据安全制度流程存在缺陷"
    assert worksheet.cell(row=2, column=12).value == risk_record_id

    worksheet.cell(row=2, column=9).value = "合规风险、技术风险"
    worksheet.cell(row=2, column=6).value = "导入后的风险描述"
    worksheet.cell(row=2, column=7).value = "导入后的风险源描述"
    worksheet.cell(row=2, column=10).value = "重要数据/二级"
    worksheet.cell(row=2, column=11).value = "收集\n存储；传输"

    imported = unwrap(
        client.post(
            f"/api/v1/projects/{project_id}/risk-sources/import",
            data={"file": (workbook_stream(workbook), "risk-sources.xlsx")},
            content_type="multipart/form-data",
        )
    )
    assert imported == {"importedCount": 1, "failedCount": 0, "errors": []}

    updated = unwrap(client.get(f"/api/v1/projects/{project_id}/risk-sources?pageNo=1&pageSize=1"))["list"][0]
    assert updated["riskTypes"] == ["合规风险", "技术风险"]
    assert updated["riskDescription"] == "导入后的风险描述"
    assert updated["riskSourceDescription"] == "导入后的风险源描述"
    assert updated["relatedData"] == "重要数据/二级"
    assert updated["relatedActivities"] == ["收集", "存储", "传输"]


def test_risk_item_excel_import_recalculates_risk_level_and_reports_duplicates(client):
    project_id, risk_record_id = seed_risk_summary(client, "ENST-TEST-RISK-ITEM-EXCEL")
    unwrap(
        client.put(
            f"/api/v1/projects/{project_id}/risk-items/{risk_record_id}",
            json={"harmLevel": "VERY_HIGH", "possibilityLevel": "LOW"},
        )
    )

    export_response = client.get(f"/api/v1/projects/{project_id}/risk-items/export")
    assert export_response.status_code == 200
    workbook = load_workbook(BytesIO(export_response.data))
    worksheet = workbook.active

    assert [cell.value for cell in worksheet[1][:10]] == [
        "序号",
        "评估项ID",
        "风险类型",
        "风险描述",
        "危害程度",
        "发生可能性",
        "风险源描述",
        "风险等级",
        "涉及的数据及类型、级别",
        "涉及的数据处理活动",
    ]
    assert worksheet.cell(row=1, column=11).value == "riskItemId"
    assert worksheet.column_dimensions["K"].hidden is True
    assert worksheet.cell(row=2, column=2).value == "AQGL001"
    assert worksheet.cell(row=2, column=5).value == "很高"
    assert worksheet.cell(row=2, column=6).value == "低"
    assert worksheet.cell(row=2, column=8).value == "中安全风险"

    worksheet.cell(row=2, column=3).value = "管理风险；技术风险"
    worksheet.cell(row=2, column=4).value = "风险清单导入描述"
    worksheet.cell(row=2, column=5).value = "很高"
    worksheet.cell(row=2, column=6).value = "低"
    worksheet.cell(row=2, column=7).value = "风险源导入描述"
    worksheet.cell(row=2, column=8).value = "MAJOR"
    worksheet.cell(row=2, column=9).value = "核心数据/三级"
    worksheet.cell(row=2, column=10).value = "传输、删除"
    worksheet.append([2, "AQGL001", "重复行", "不应覆盖", "中", "高", "重复源", "LOW", "重复数据", "收集", risk_record_id])

    imported = unwrap(
        client.post(
            f"/api/v1/projects/{project_id}/risk-items/import",
            data={"file": (workbook_stream(workbook), "risk-items.xlsx")},
            content_type="multipart/form-data",
        )
    )
    assert imported["importedCount"] == 1
    assert imported["failedCount"] == 1
    assert imported["errors"] == [
        {
            "rowNo": 3,
            "field": "riskItemId",
            "reason": "同一导入文件中已出现相同风险记录。",
        }
    ]

    risk_item = unwrap(client.get(f"/api/v1/projects/{project_id}/risk-items?pageNo=1&pageSize=1"))["list"][0]
    assert risk_item["riskTypes"] == ["管理风险", "技术风险"]
    assert risk_item["riskDescription"] == "风险清单导入描述"
    assert risk_item["harmLevel"] == "VERY_HIGH"
    assert risk_item["possibilityLevel"] == "LOW"
    assert risk_item["riskSourceDescription"] == "风险源导入描述"
    assert risk_item["riskLevel"] == "MEDIUM"
    assert risk_item["relatedData"] == "核心数据/三级"
    assert risk_item["relatedActivities"] == ["传输", "删除"]

    session = SessionLocal()
    records = session.query(ProjectRiskSummaryRecord).filter_by(project_id=project_id, deleted=False).all()
    assert len(records) == 1


def test_risk_item_import_without_locator_returns_row_error(client):
    project_id, _risk_record_id = seed_risk_summary(client, "ENST-TEST-RISK-ITEM-NO-LOCATOR")

    export_response = client.get(f"/api/v1/projects/{project_id}/risk-items/export")
    workbook = load_workbook(BytesIO(export_response.data))
    worksheet = workbook.active
    worksheet.cell(row=2, column=4).value = "不应在无法定位时写入"
    worksheet.cell(row=2, column=11).value = None

    imported = unwrap(
        client.post(
            f"/api/v1/projects/{project_id}/risk-items/import",
            data={"file": (workbook_stream(workbook), "risk-items.xlsx")},
            content_type="multipart/form-data",
        )
    )
    assert imported == {
        "importedCount": 0,
        "failedCount": 1,
        "errors": [
            {
                "rowNo": 2,
                "field": "riskItemId",
                "reason": "无法定位对应风险记录，请使用最新导出模板。",
            }
        ],
    }

    risk_item = unwrap(client.get(f"/api/v1/projects/{project_id}/risk-items?pageNo=1&pageSize=1"))["list"][0]
    assert risk_item["riskDescription"] != "不应在无法定位时写入"
