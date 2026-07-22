from io import BytesIO
from pathlib import Path
from xml.etree import ElementTree as ET
import zipfile

from openpyxl import Workbook, load_workbook

from app.extensions import SessionLocal
from app.models import (
    AssessedOrganization,
    AssessmentTeamMember,
    AssessmentTemplateItem,
    BusinessSystem,
    ClientTeamMember,
    DataProcessorBasicSurvey,
    EvaluationRecord,
    FocusPoint,
    GapItem,
    ProcessingActivitySurvey,
    ProjectBasicInfo,
    RiskSourceTemplate,
    SecurityProtectionSurvey,
)


def unwrap(response):
    assert response.status_code < 400, response.get_data(as_text=True)
    payload = response.get_json()
    assert payload["code"] == "SUCCESS"
    return payload["data"]


def create_project(client, score_model_id="score-v1", project_code="ENST-TEST-001", risk_matrix_id="matrix-v1"):
    return unwrap(
        client.post(
            "/api/v1/projects",
            json={
                "projectName": "Flow Project",
                "projectCode": project_code,
                "assessmentOrg": "Test Org",
                "assessmentTemplateId": "tpl-gb",
                "scoreModelId": score_model_id,
                "harmModelId": "harm-default",
                "riskMatrixId": risk_matrix_id,
                "systemType": "MANAGEMENT_SYSTEM",
            },
        )
    )["projectId"]


def workbook_stream(workbook):
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def survey_docx_stream(edits):
    template = Path("doc") / "附录A（资料性）调研表格.docx"
    source = BytesIO(template.read_bytes())
    output = BytesIO()
    namespace = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
    with zipfile.ZipFile(source) as zin, zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as zout:
        document = ET.fromstring(zin.read("word/document.xml"))
        tables = list(document.iter(namespace + "tbl"))
        for table_no, row_no, col_no, value in edits:
            cell = tables[table_no].findall(namespace + "tr")[row_no].findall(namespace + "tc")[col_no]
            texts = list(cell.iter(namespace + "t"))
            if texts:
                texts[0].text = value
                for extra in texts[1:]:
                    extra.text = ""
            else:
                paragraph = cell.find(namespace + "p")
                if paragraph is None:
                    paragraph = ET.SubElement(cell, namespace + "p")
                run = ET.SubElement(paragraph, namespace + "r")
                text = ET.SubElement(run, namespace + "t")
                text.text = value
        xml = ET.tostring(document, encoding="utf-8", xml_declaration=True)
        for item in zin.infolist():
            zout.writestr(item, xml if item.filename == "word/document.xml" else zin.read(item.filename))
    output.seek(0)
    return output


def survey_docx_power_decomposed_stream(power_rows, threat_value=""):
    template = Path("doc") / "附录A（资料性）调研表格.docx"
    source = BytesIO(template.read_bytes())
    output = BytesIO()
    namespace = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
    with zipfile.ZipFile(source) as zin, zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as zout:
        document = ET.fromstring(zin.read("word/document.xml"))
        tables = list(document.iter(namespace + "tbl"))
        security_table = tables[7]
        rows = security_table.findall(namespace + "tr")
        row9 = _docx_sequence_row(security_table, namespace, "9")
        row10 = _docx_sequence_row(security_table, namespace, "10")
        row9_child_index = list(security_table).index(row9)
        row10_child_index = list(security_table).index(row10) if row10 is not None else len(list(security_table))
        for row in list(security_table)[row9_child_index:row10_child_index]:
            security_table.remove(row)
        for offset, values in enumerate(power_rows):
            security_table.insert(row9_child_index + offset, _docx_row(values, namespace))
        if threat_value:
            row10 = _docx_sequence_row(security_table, namespace, "10")
            _set_docx_cell_text(row10.findall(namespace + "tc")[-1], threat_value, namespace)
        xml = ET.tostring(document, encoding="utf-8", xml_declaration=True)
        for item in zin.infolist():
            zout.writestr(item, xml if item.filename == "word/document.xml" else zin.read(item.filename))
    output.seek(0)
    return output


def _docx_sequence_row(table, namespace, sequence):
    for row in table.findall(namespace + "tr"):
        cells = row.findall(namespace + "tc")
        if cells and "".join(t.text or "" for t in cells[0].iter(namespace + "t")).strip() == sequence:
            return row
    return None


def _docx_row(values, namespace):
    row = ET.Element(namespace + "tr")
    for value in values:
        cell = ET.Element(namespace + "tc")
        paragraph = ET.SubElement(cell, namespace + "p")
        run = ET.SubElement(paragraph, namespace + "r")
        text = ET.SubElement(run, namespace + "t")
        text.text = value
        row.append(cell)
    return row


def _set_docx_cell_text(cell, value, namespace):
    texts = list(cell.iter(namespace + "t"))
    if texts:
        texts[0].text = value
        for extra in texts[1:]:
            extra.text = ""
        return
    paragraph = cell.find(namespace + "p")
    if paragraph is None:
        paragraph = ET.SubElement(cell, namespace + "p")
    run = ET.SubElement(paragraph, namespace + "r")
    text = ET.SubElement(run, namespace + "t")
    text.text = value


def test_project_create_start_and_list(client):
    project_id = create_project(client)
    assert unwrap(client.get("/api/v1/projects/statistics"))["notStarted"] == 1

    started = unwrap(client.post(f"/api/v1/projects/{project_id}/start"))
    assert started["status"] == "IN_PROGRESS"
    assert started["generatedItemCount"] > 0

    restarted = unwrap(client.post(f"/api/v1/projects/{project_id}/start"))
    assert restarted["generatedItemCount"] == 0
    assert restarted["existingItemCount"] > 0

    projects = unwrap(client.get("/api/v1/projects?pageNo=1&pageSize=10"))
    assert projects["total"] == 1


def test_project_code_can_be_shared_by_multiple_system_projects(client):
    first_id = create_project(client, project_code="ENST-SHARED-001")
    second_id = create_project(client, project_code="ENST-SHARED-001")

    assert first_id != second_id
    projects = unwrap(client.get("/api/v1/projects?pageNo=1&pageSize=10&keyword=ENST-SHARED-001"))
    assert projects["total"] == 2
    assert {project["projectCode"] for project in projects["list"]} == {"ENST-SHARED-001"}


def test_project_code_can_be_updated_to_existing_code(client):
    first_id = create_project(client, project_code="ENST-SHARED-EDIT-001")
    second_id = create_project(client, project_code="ENST-SHARED-EDIT-002")

    updated = unwrap(
        client.put(
            f"/api/v1/projects/{second_id}",
            json={
                "projectName": "Flow Project",
                "projectCode": "ENST-SHARED-EDIT-001",
                "assessmentOrg": "Test Org",
            },
        )
    )

    assert first_id != second_id
    assert updated["projectCode"] == "ENST-SHARED-EDIT-001"
    projects = unwrap(client.get("/api/v1/projects?pageNo=1&pageSize=10&keyword=ENST-SHARED-EDIT-001"))
    assert projects["total"] == 2


def test_assessment_item_id_flows_to_evaluation_and_risk_lists(client):
    session = SessionLocal()
    template_items = (
        session.query(AssessmentTemplateItem)
        .filter(AssessmentTemplateItem.template_id == "tpl-gb", AssessmentTemplateItem.deleted.is_(False))
        .all()
    )
    assert len(template_items) == 391
    assert not any(item.sheet_name.endswith(("113-3-20", "125", "76+13", "87")) for item in template_items)
    assert any(item.assessment_item_id == "AQJS077" and item.sheet_name == "数据安全技术" for item in template_items)
    risk_template = (
        session.query(RiskSourceTemplate)
        .filter(
            RiskSourceTemplate.sheet_name == "数据安全管理",
            RiskSourceTemplate.category == "安全管理制度",
            RiskSourceTemplate.subcategory == "数据安全制度体系",
            RiskSourceTemplate.deleted.is_(False),
        )
        .order_by(RiskSourceTemplate.sort_order.asc())
        .first()
    )
    risk_template.risk_source_type = "数据安全制度流程存在缺陷"
    session.commit()

    project_id = create_project(client, project_code="ENST-TEST-ASSESSMENT-ID")
    unwrap(client.post(f"/api/v1/projects/{project_id}/start"))

    items = unwrap(client.get(f"/api/v1/projects/{project_id}/evaluation/items?pageNo=1&pageSize=1"))
    evaluation_item = items["list"][0]
    assert evaluation_item["assessmentItemId"] == "AQGL001"
    assert evaluation_item["sheetName"] == "数据安全管理"

    unwrap(
        client.put(
            f"/api/v1/projects/{project_id}/evaluation/items/{evaluation_item['id']}/record",
            json={"evaluationResult": "PARTIAL", "evaluationRecord": "assessment id risk record"},
        )
    )
    unwrap(client.post(f"/api/v1/projects/{project_id}/risk-summary/refresh", json={}))

    for path in ["risk-sources", "risk-items", "risk-suggestions"]:
        page = unwrap(client.get(f"/api/v1/projects/{project_id}/{path}?pageNo=1&pageSize=1"))
        assert page["list"][0]["assessmentItemId"] == "AQGL001"
    source_page = unwrap(client.get(f"/api/v1/projects/{project_id}/risk-sources?pageNo=1&pageSize=1"))
    assert source_page["list"][0]["assessmentSubcategory"] == "数据安全制度体系"
    assert "assessmentSubitem" not in source_page["list"][0]
    assert source_page["list"][0]["riskSourceType"] == "数据安全制度流程存在缺陷"


def test_default_risk_matrix_matches_linkage_chart(client):
    from app.services import harm_analysis_service
    from app.services.project_service import get_project

    project_id = create_project(client, project_code="ENST-TEST-MATRIX")
    project = get_project(project_id)
    session = SessionLocal()
    expected = {
        "HIGH": {
            "VERY_HIGH": "MAJOR",
            "HIGH": "MAJOR",
            "RELATIVELY_HIGH": "MEDIUM",
            "MEDIUM": "LOW",
            "LOW": "SLIGHT",
        },
        "MEDIUM": {
            "VERY_HIGH": "MAJOR",
            "HIGH": "HIGH",
            "RELATIVELY_HIGH": "MEDIUM",
            "MEDIUM": "LOW",
            "LOW": "SLIGHT",
        },
        "LOW": {
            "VERY_HIGH": "MEDIUM",
            "HIGH": "MEDIUM",
            "RELATIVELY_HIGH": "LOW",
            "MEDIUM": "SLIGHT",
            "LOW": "SLIGHT",
        },
    }
    for possibility_level, row in expected.items():
        for harm_level, risk_level in row.items():
            assert harm_analysis_service.risk_level_from_project_matrix(session, project, harm_level, possibility_level) == risk_level


def test_risk_level_uses_project_selected_knowledge_matrix(client):
    from app.services import harm_analysis_service
    from app.services.project_service import get_project

    custom_matrix = unwrap(
        client.post(
            "/api/v1/knowledge/risk-matrices",
            json={
                "matrixName": "项目自定义风险评价矩阵",
                "matrixJson": {
                    "HIGH": {
                        "VERY_HIGH": "MAJOR",
                        "HIGH": "MAJOR",
                        "RELATIVELY_HIGH": "MAJOR",
                        "MEDIUM": "HIGH",
                        "LOW": "MEDIUM",
                    },
                    "MEDIUM": {
                        "VERY_HIGH": "HIGH",
                        "HIGH": "HIGH",
                        "RELATIVELY_HIGH": "HIGH",
                        "MEDIUM": "MEDIUM",
                        "LOW": "LOW",
                    },
                    "LOW": {
                        "VERY_HIGH": "MEDIUM",
                        "HIGH": "MEDIUM",
                        "RELATIVELY_HIGH": "MEDIUM",
                        "MEDIUM": "LOW",
                        "LOW": "SLIGHT",
                    },
                },
            },
        )
    )
    project_id = create_project(client, project_code="ENST-TEST-CUSTOM-MATRIX", risk_matrix_id=custom_matrix["id"])
    project = get_project(project_id)
    session = SessionLocal()

    assert harm_analysis_service.risk_level_from_project_matrix(session, project, "RELATIVELY_HIGH", "HIGH") == "MAJOR"


def test_score_uses_selected_score_model_result_values(client):
    score_model = unwrap(
        client.post(
            "/api/v1/knowledge/score-models",
            json={
                "modelName": "Custom Result Value Model",
                "modelType": "CUSTOM",
                "resultScores": {
                    "COMPLIANT": 0.8,
                    "PARTIAL": 0.25,
                    "NON_COMPLIANT": 0.1,
                    "NOT_APPLICABLE": None,
                },
                "possibilityRanges": [
                    {"level": "HIGH", "min": 0, "max": 60, "includeMin": True, "includeMax": False},
                    {"level": "MEDIUM", "min": 60, "max": 80, "includeMin": True, "includeMax": False},
                    {"level": "LOW", "min": 80, "max": 100, "includeMin": True, "includeMax": True},
                ],
            },
        )
    )
    project_id = create_project(client, score_model_id=score_model["id"], project_code="ENST-TEST-PARTIAL")
    unwrap(client.post(f"/api/v1/projects/{project_id}/start"))

    items = unwrap(client.get(f"/api/v1/projects/{project_id}/evaluation/items?pageNo=1&pageSize=3"))
    for item, result in zip(items["list"], ["COMPLIANT", "PARTIAL", "NON_COMPLIANT"]):
        unwrap(
            client.put(
                f"/api/v1/projects/{project_id}/evaluation/items/{item['id']}/record",
                json={"evaluationResult": result, "evaluationRecord": f"{result} record"},
            )
        )

    score = unwrap(client.post(f"/api/v1/projects/{project_id}/evaluation/calculate-score"))
    assert score["score"] == 38.33
    assert score["detail"]["compliantCount"] == 1
    assert score["detail"]["partialCount"] == 1
    assert score["detail"]["nonCompliantCount"] == 1
    assert score["scoreModelVersion"] == score_model["version"]


def test_evaluation_items_support_multi_result_filters(client):
    project_id = create_project(client, project_code="ENST-TEST-EVALUATION-FILTERS")
    unwrap(client.post(f"/api/v1/projects/{project_id}/start"))

    items = unwrap(client.get(f"/api/v1/projects/{project_id}/evaluation/items?pageNo=1&pageSize=3"))["list"]
    for item, result in zip(items, ["COMPLIANT", "PARTIAL", "NON_COMPLIANT"]):
        unwrap(
            client.put(
                f"/api/v1/projects/{project_id}/evaluation/items/{item['id']}/record",
                json={"evaluationResult": result, "evaluationRecord": f"{result} record"},
            )
        )

    multi = unwrap(
        client.get(
            f"/api/v1/projects/{project_id}/evaluation/items",
            query_string=[
                ("pageNo", "1"),
                ("pageSize", "10"),
                ("results[]", "PARTIAL"),
                ("results[]", "NON_COMPLIANT"),
            ],
        )
    )
    assert [item["evaluationResult"] for item in multi["list"]] == ["PARTIAL", "NON_COMPLIANT"]
    assert multi["total"] == 2

    repeated = unwrap(
        client.get(
            f"/api/v1/projects/{project_id}/evaluation/items",
            query_string=[
                ("pageNo", "1"),
                ("pageSize", "10"),
                ("results", "PARTIAL"),
                ("results", "NON_COMPLIANT"),
            ],
        )
    )
    assert [item["evaluationResult"] for item in repeated["list"]] == ["PARTIAL", "NON_COMPLIANT"]

    comma_with_keyword = unwrap(
        client.get(
            f"/api/v1/projects/{project_id}/evaluation/items",
            query_string={
                "pageNo": "1",
                "pageSize": "10",
                "results": "PARTIAL,NON_COMPLIANT",
                "keyword": items[1]["itemCode"],
            },
        )
    )
    assert comma_with_keyword["total"] == 1
    assert comma_with_keyword["list"][0]["id"] == items[1]["id"]
    assert comma_with_keyword["list"][0]["evaluationResult"] == "PARTIAL"

    sheet_and_category = unwrap(
        client.get(
            f"/api/v1/projects/{project_id}/evaluation/items",
            query_string={
                "pageNo": "1",
                "pageSize": "10",
                "sheetName": items[1]["sheetName"],
                "category": items[1]["category"],
                "result": "PARTIAL",
            },
        )
    )
    assert sheet_and_category["total"] == 1
    assert sheet_and_category["list"][0]["id"] == items[1]["id"]

    legacy = unwrap(
        client.get(
            f"/api/v1/projects/{project_id}/evaluation/items",
            query_string={"pageNo": "1", "pageSize": "10", "result": "PARTIAL"},
        )
    )
    assert [item["evaluationResult"] for item in legacy["list"]] == ["PARTIAL"]

    precedence = unwrap(
        client.get(
            f"/api/v1/projects/{project_id}/evaluation/items",
            query_string=[
                ("pageNo", "1"),
                ("pageSize", "10"),
                ("results[]", "NON_COMPLIANT"),
                ("result", "PARTIAL"),
            ],
        )
    )
    assert [item["evaluationResult"] for item in precedence["list"]] == ["NON_COMPLIANT"]

    invalid = client.get(
        f"/api/v1/projects/{project_id}/evaluation/items",
        query_string={"pageNo": "1", "pageSize": "10", "results[]": "UNKNOWN"},
    )
    assert invalid.status_code == 400
    payload = invalid.get_json()
    assert payload["code"] == "INVALID_EVALUATION_RESULT_FILTER"
    assert payload["message"] == "符合情况筛选值只能为 COMPLIANT、PARTIAL、NON_COMPLIANT、NOT_APPLICABLE"


def test_risk_refresh_uses_score_model_boundary_for_possibility(client):
    score_model = unwrap(
        client.post(
            "/api/v1/knowledge/score-models",
            json={
                "modelName": "Boundary Model",
                "modelType": "CUSTOM",
                "resultScores": {
                    "COMPLIANT": 1,
                    "PARTIAL": 0.5,
                    "NON_COMPLIANT": 0,
                    "NOT_APPLICABLE": None,
                },
                "possibilityRanges": [
                    {"level": "HIGH", "min": 0, "max": 30, "includeMin": True, "includeMax": False},
                    {"level": "MEDIUM", "min": 30, "max": 75, "includeMin": True, "includeMax": False},
                    {"level": "LOW", "min": 75, "max": 100, "includeMin": True, "includeMax": True},
                ],
            },
        )
    )
    project_id = create_project(client, score_model_id=score_model["id"], project_code="ENST-TEST-BOUNDARY")
    unwrap(client.post(f"/api/v1/projects/{project_id}/start"))

    items = unwrap(client.get(f"/api/v1/projects/{project_id}/evaluation/items?pageNo=1&pageSize=2"))
    unwrap(
        client.put(
            f"/api/v1/projects/{project_id}/evaluation/items/{items['list'][0]['id']}/record",
            json={"evaluationResult": "COMPLIANT", "evaluationRecord": "compliant"},
        )
    )
    unwrap(
        client.put(
            f"/api/v1/projects/{project_id}/evaluation/items/{items['list'][1]['id']}/record",
            json={"evaluationResult": "PARTIAL", "evaluationRecord": "partial"},
        )
    )

    score = unwrap(client.post(f"/api/v1/projects/{project_id}/evaluation/calculate-score"))
    assert score["score"] == 75.0
    assert score["possibilityLevel"] == "LOW"

    refreshed = unwrap(client.post(f"/api/v1/projects/{project_id}/risk-summary/refresh", json={}))
    assert refreshed["score"] == 75.0
    assert refreshed["possibilityLevel"] == "LOW"

    sources = unwrap(client.get(f"/api/v1/projects/{project_id}/risk-sources?pageNo=1&pageSize=1"))
    assert sources["pageNo"] == 1
    assert sources["pageSize"] == 1
    assert sources["total"] == 1
    assert sources["list"][0]["possibilityLevel"] == "LOW"


def test_basic_plan_survey_evaluation_and_risk_flow(client):
    project_id = create_project(client)
    unwrap(client.post(f"/api/v1/projects/{project_id}/start"))

    basic = unwrap(
        client.put(
            f"/api/v1/projects/{project_id}/basic-info",
            json={
                "projectNumber": "SHOULD-NOT-CHANGE",
                "projectName": "新版项目名称",
                "laws": [{"id": "law-1", "name": "law"}],
                "standards": [{"id": "std-1", "name": "std"}],
                "assessmentPlan": {"startDate": "2026-06-01", "endDate": "2026-06-30"},
                "organization": {"name": "org", "postalCode": "100000"},
                "contacts": [{"name": "alice", "email": "alice@example.com"}],
            },
        )
    )
    assert basic["projectNumber"] == "ENST-TEST-001"
    assert basic["projectName"] == "新版项目名称"
    assert basic["organization"]["name"] == "org"
    assert basic["organization"]["postalCode"] == "100000"
    assert "projectDescription" not in basic
    assert "systemDescription" not in basic
    assert "assessmentTarget" not in basic
    assert "creditCode" not in basic["organization"]
    assert "project_description" not in ProjectBasicInfo.__table__.columns
    assert "system_description" not in ProjectBasicInfo.__table__.columns
    assert "assessment_target" not in ProjectBasicInfo.__table__.columns
    assert "address" not in AssessedOrganization.__table__.columns
    assert "credit_code" not in AssessedOrganization.__table__.columns
    assert "data_security_owner" not in AssessedOrganization.__table__.columns
    assert "description" not in AssessedOrganization.__table__.columns

    team = unwrap(
        client.post(
            f"/api/v1/projects/{project_id}/plan/assessment-team",
            json={"name": "Alice", "organization": "ENST", "role": "leader"},
        )
    )
    assert team["name"] == "Alice"

    asset = unwrap(
        client.post(
            f"/api/v1/projects/{project_id}/survey/data-assets",
            json={"dataName": "Customer Data", "dataLevel": "important", "personalInfo": True},
        )
    )
    assert asset["dataName"] == "Customer Data"

    personal = unwrap(
        client.post(
            f"/api/v1/projects/{project_id}/survey/personal-info",
            json={
                "dataName": "用户手机号",
                "dataCategory": "个人信息-联系方式",
                "scale": "10万条",
                "dataSource": "用户注册",
                "businessFlow": "注册后进入客户管理系统",
                "sensitivity": "敏感",
            },
        )
    )
    assert personal["dataName"] == "用户手机号"
    assert personal["businessFlow"] == "注册后进入客户管理系统"

    processor_basic = unwrap(
        client.put(
            f"/api/v1/projects/{project_id}/survey/data-processor-basic",
            json={
                "unitName": "数据处理者",
                "unitNature": "企业",
                "unifiedSocialCreditCode": "91330000000000000X",
                "businessScope": "电力业务",
                "isIgnoredByBackend": "ignored",
            },
        )
    )
    assert processor_basic["unitName"] == "数据处理者"
    assert processor_basic["unitNature"] == "企业"
    assert "isIgnoredByBackend" not in processor_basic
    assert unwrap(client.get(f"/api/v1/projects/{project_id}/survey/data-processor-basic"))["businessScope"] == "电力业务"
    session = SessionLocal()
    processor_basic_row = session.query(DataProcessorBasicSurvey).filter_by(project_id=project_id).one()
    assert "payload" not in DataProcessorBasicSurvey.__table__.columns
    assert processor_basic_row.unit_name == "数据处理者"
    assert processor_basic_row.business_scope == "电力业务"

    business_system = unwrap(
        client.post(
            f"/api/v1/projects/{project_id}/survey/business-systems",
            json={
                "systemName": "核心业务系统",
                "businessFunction": "业务办理",
                "dataScopes": ["一般数据", "个人信息", "重要数据"],
            },
        )
    )
    assert business_system["dataScopes"] == ["一般数据", "个人信息", "重要数据"]
    business_system_row = session.query(BusinessSystem).filter_by(id=business_system["id"]).one()
    assert business_system_row.data_scopes == "一般数据、个人信息、重要数据"

    activity_survey_default = unwrap(client.get(f"/api/v1/projects/{project_id}/survey/processing-activity-survey"))
    assert activity_survey_default["involvedActivities"] == []
    assert activity_survey_default["collectionChannels"] == ""
    activity_survey = unwrap(
        client.put(
            f"/api/v1/projects/{project_id}/survey/processing-activity-survey",
            json={
                "involvedActivities": ["COLLECT", "TRANSFER"],
                "collectionChannels": "线上渠道",
                "collectionMethod": "接口采集",
                "transferProtocol": "HTTPS",
                "storageMethod": "不应保存",
                "collect": {"scenarios": ["旧字段不应返回"]},
            },
        )
    )
    assert activity_survey["involvedActivities"] == ["COLLECT", "TRANSFER"]
    assert "activityTypes" not in activity_survey
    assert "collect" not in activity_survey
    assert activity_survey["collectionChannels"] == "线上渠道"
    assert activity_survey["collectionMethod"] == "接口采集"
    assert activity_survey["transferProtocol"] == "HTTPS"
    assert activity_survey["storageMethod"] == "不应保存"
    activity_survey_row = session.query(ProcessingActivitySurvey).filter_by(project_id=project_id).one()
    assert "payload" not in ProcessingActivitySurvey.__table__.columns
    assert activity_survey_row.involved_activities == "COLLECT,TRANSFER"
    assert activity_survey_row.collection_channels == "线上渠道"
    assert activity_survey_row.collection_method == "接口采集"
    assert activity_survey_row.transfer_protocol == "HTTPS"
    assert activity_survey_row.storage_method == "不应保存"

    protection = unwrap(
        client.put(
            f"/api/v1/projects/{project_id}/survey/security-protection",
            json={
                "complianceAssessmentStatus": "已完成等保测评",
                "identityAuthenticationAndAccessControl": "密码和多因素认证",
                "isPowerMonitoringSystem": "否",
                "productionControlAreaProtection": "旧电力监控字段可独立保存",
                "classifiedProtectionDone": "NO",
            },
        )
    )
    assert protection["complianceAssessmentStatus"] == "已完成等保测评"
    assert protection["identityAuthenticationAndAccessControl"] == "密码和多因素认证"
    assert protection["isPowerMonitoringSystem"] == "NO"
    assert protection["productionControlAreaProtection"] == "旧电力监控字段可独立保存"
    assert "classifiedProtectionDone" not in protection
    protection_row = session.query(SecurityProtectionSurvey).filter_by(project_id=project_id).one()
    assert "payload" not in SecurityProtectionSurvey.__table__.columns
    assert protection_row.compliance_assessment_status == "已完成等保测评"
    assert protection_row.identity_authentication_and_access_control == "密码和多因素认证"
    assert protection_row.is_power_monitoring_system == "NO"

    protection_yes = unwrap(
        client.put(
            f"/api/v1/projects/{project_id}/survey/security-protection",
            json={
                "isPowerMonitoringSystem": "YES",
                "powerDispatchAuthentication": "已部署调度数字证书",
            },
        )
    )
    assert protection_yes["isPowerMonitoringSystem"] == "YES"
    assert protection_yes["powerDispatchAuthentication"] == "已部署调度数字证书"

    items = unwrap(client.get(f"/api/v1/projects/{project_id}/evaluation/items?pageNo=1&pageSize=1"))
    item_id = items["list"][0]["id"]
    saved_record = unwrap(
        client.put(
            f"/api/v1/projects/{project_id}/evaluation/items/{item_id}/record",
            json={
                "evaluationResult": "PARTIAL",
                "evaluationRecord": "record",
            },
        )
    )
    assert saved_record["evaluationRecord"] == "record"
    assert "problemDescription" not in saved_record
    assert "riskTypes" not in saved_record
    score = unwrap(client.post(f"/api/v1/projects/{project_id}/evaluation/calculate-score"))
    assert score["detail"]["partialCount"] == 1

    refreshed = unwrap(client.post(f"/api/v1/projects/{project_id}/risk-summary/refresh", json={}))
    assert refreshed["createdSources"] == 1

    sources = unwrap(client.get(f"/api/v1/projects/{project_id}/risk-sources?pageNo=1&pageSize=10"))
    assert sources["total"] == 1
    source_id = sources["list"][0]["id"]
    assert sources["list"][0]["riskRecordId"] == source_id
    assert sources["list"][0]["checkPoint"]
    assert sources["list"][0]["evaluationRecord"] == "record"
    assert sources["list"][0]["assessmentCategory"]
    assert sources["list"][0]["assessmentSubcategory"]
    assert sources["list"][0]["riskDescription"]
    assert sources["list"][0]["riskTypes"]
    assert sources["list"][0]["relatedData"] == "-"
    assert sources["list"][0]["harmLevel"] == "-"
    assert sources["list"][0]["possibilityLevel"] == "HIGH"
    assert sources["list"][0]["riskLevel"] is None
    risk_items = unwrap(client.get(f"/api/v1/projects/{project_id}/risk-items?pageNo=1&pageSize=1"))
    assert risk_items["pageSize"] == 1
    assert risk_items["total"] == 1
    risk_suggestions = unwrap(client.get(f"/api/v1/projects/{project_id}/risk-suggestions?pageNo=1&pageSize=1"))
    assert risk_suggestions["pageSize"] == 1
    assert risk_suggestions["total"] == 1

    llm_result = {
        "impactedObject": "PUBLIC_INTEREST",
        "damageDegree": "SERIOUS",
        "reason": "影响电力供应连续性，可能对公共利益造成严重危害。",
        "confidence": 0.82,
    }
    suggestion = unwrap(
        client.post(
            f"/api/v1/projects/{project_id}/risk-items/{source_id}/harm-analysis/suggest",
            json={"llmResult": llm_result},
        )
    )
    assert suggestion["harmLevel"] == "RELATIVELY_HIGH"
    assert suggestion["dataSecurityProtectionLevel"] == 3
    assert suggestion["harmImpactObject"] == "社会秩序和公共利益"
    assert "影响电力供应连续性" in suggestion["harmDescription"]
    assert suggestion["needsManualReview"] is False
    assert suggestion["harmAnalysisTrace"]["step3"]
    assert suggestion["harmAnalysisTrace"]["step4"] == "按保护等级匹配风险危害程度等级：较高。"
    assert suggestion["harmAnalysisTrace"]["step2Reason"] == "影响电力供应连续性，可能对公共利益造成严重危害。"
    assert "符合情况：基本符合" in suggestion["harmAnalysisTrace"]["step2Basis"]
    assert "符合情况：PARTIAL" not in suggestion["harmAnalysisTrace"]["step2Basis"]
    assert any("项目系统类别" in item for item in suggestion["harmAnalysisTrace"]["step2Basis"])
    assert any("判定规则" in item for item in suggestion["harmAnalysisTrace"]["step2Basis"])
    assert suggestion["harmAnalysisTrace"]["step2Evidence"] == []

    unchanged = unwrap(client.get(f"/api/v1/projects/{project_id}/risk-items?pageNo=1&pageSize=1"))
    assert unchanged["list"][0]["harmLevel"] == "-"
    assert unchanged["list"][0]["harmAnalysisStatus"] == "PENDING"

    batch = unwrap(
        client.post(
            f"/api/v1/projects/{project_id}/risk-items/harm-analysis/suggest-batch",
            json={"riskItemIds": [source_id], "llmResult": llm_result},
        )
    )
    assert batch["total"] == 1
    assert batch["suggestions"][0]["riskRecordId"] == source_id

    applied_harm = unwrap(
        client.post(
            f"/api/v1/projects/{project_id}/risk-items/{source_id}/harm-analysis/apply",
            json={"suggestion": suggestion},
        )
    )
    assert applied_harm["harmLevel"] == "RELATIVELY_HIGH"
    assert applied_harm["harmDescription"]
    assert applied_harm["harmExample"]
    assert applied_harm["harmAnalysisTrace"]["dataSecurityProtectionLevel"] == 3
    assert applied_harm["harmAnalysisStatus"] == "CONFIRMED"
    assert applied_harm["riskLevel"] == "MEDIUM"

    original_risk_description = sources["list"][0]["riskDescription"]
    original_source_description = sources["list"][0]["riskSourceDescription"]
    updated_item = unwrap(
        client.put(
            f"/api/v1/projects/{project_id}/risk-items/{source_id}",
            json={
                "riskTypes": ["IGNORED"],
                "riskDescription": "ignored risk",
                "riskSourceDescription": "ignored source",
                "relatedData": "ignored data",
                "relatedActivities": ["IGNORED"],
                "harmLevel": "MEDIUM",
                "riskLevel": "MAJOR",
            },
        )
    )
    assert updated_item["riskDescription"] == original_risk_description
    assert updated_item["riskSourceDescription"] == original_source_description
    assert updated_item["relatedData"] == "-"
    assert updated_item["relatedActivities"] == []
    assert updated_item["harmLevel"] == "MEDIUM"
    assert updated_item["riskLevel"] == "LOW"

    updated_suggestion = unwrap(
        client.put(
            f"/api/v1/projects/{project_id}/risk-suggestions/{source_id}",
            json={"riskDescription": "ignored", "riskSourceDescription": "ignored", "riskLevel": "HIGH", "remediationSuggestion": "manual suggestion"},
        )
    )
    assert updated_suggestion["riskDescription"] == original_risk_description
    assert updated_suggestion["riskSourceDescription"] == original_source_description
    assert updated_suggestion["riskLevel"] == "LOW"
    assert updated_suggestion["remediationSuggestion"] == "manual suggestion"

    updated = unwrap(
        client.put(
            f"/api/v1/projects/{project_id}/risk-sources/{source_id}",
            json={"riskSourceDescription": "manual problem"},
        )
    )
    assert updated["riskSourceDescription"] == "manual problem"


def test_basic_info_excel_template_import_and_export(client):
    project_id = create_project(client, project_code="ENST-TEST-BASIC-EXCEL")

    template_response = client.get(f"/api/v1/projects/{project_id}/basic-info/export-template")
    assert template_response.status_code == 200
    assert template_response.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    workbook = load_workbook(BytesIO(template_response.data))
    assert workbook.sheetnames == ["项目基本情况", "被评估单位基本信息", "联系人信息"]
    assert [cell.value for cell in workbook["项目基本情况"][1]] == [
        "项目编号",
        "项目名称",
        "评估所依据的法律法规",
        "评估所参考的标准规范",
        "评估开始日期",
        "评估结束日期",
    ]
    assert [cell.value for cell in workbook["被评估单位基本信息"][1]] == [
        "单位名称",
        "邮政编码",
    ]
    assert [cell.value for cell in workbook["联系人信息"][1]] == [
        "姓名",
        "所属部门",
        "移动电话",
        "职务/职称",
        "办公电话",
        "电子邮件",
    ]

    workbook["项目基本情况"].append([
        "ENST-TEST-BASIC-EXCEL",
        "导入后的项目名称",
        "网络安全法、数据安全法",
        "GB/T 35273; GB/T 22239",
        "2026-06-01",
        "2026-06-10",
    ])
    workbook["被评估单位基本信息"].append([
        "被评估单位",
        "310000",
    ])
    workbook["联系人信息"].append([
        "张三",
        "信息部",
        "13800000000",
        "经理",
        "0571-88888888",
        "zhangsan@example.com",
    ])

    imported = unwrap(
        client.post(
            f"/api/v1/projects/{project_id}/basic-info/import",
            data={"file": (workbook_stream(workbook), "basic-info.xlsx")},
            content_type="multipart/form-data",
        )
    )
    assert imported["projectNumber"] == "ENST-TEST-BASIC-EXCEL"
    assert imported["projectName"] == "导入后的项目名称"
    assert imported["laws"] == [{"id": "网络安全法", "name": "网络安全法"}, {"id": "数据安全法", "name": "数据安全法"}]
    assert imported["standards"] == [{"id": "GB/T 35273", "name": "GB/T 35273"}, {"id": "GB/T 22239", "name": "GB/T 22239"}]
    assert imported["assessmentPlan"] == {"startDate": "2026-06-01", "endDate": "2026-06-10"}
    assert imported["organization"]["postalCode"] == "310000"
    assert "assessmentTarget" not in imported
    assert "creditCode" not in imported["organization"]
    assert imported["contacts"][0]["email"] == "zhangsan@example.com"

    not_persisted = unwrap(client.get(f"/api/v1/projects/{project_id}/basic-info"))
    assert not_persisted["projectName"] == "Flow Project"
    assert not_persisted["contacts"] == []

    saved = unwrap(
        client.put(
            f"/api/v1/projects/{project_id}/basic-info",
            json=imported,
        )
    )
    assert saved["projectName"] == "导入后的项目名称"

    export_response = client.get(f"/api/v1/projects/{project_id}/basic-info/export")
    assert export_response.status_code == 200
    exported = load_workbook(BytesIO(export_response.data))
    assert exported.sheetnames == ["项目基本情况", "被评估单位基本信息", "联系人信息"]
    assert exported["项目基本情况"].cell(row=2, column=1).value == "ENST-TEST-BASIC-EXCEL"
    assert exported["项目基本情况"].cell(row=2, column=2).value == "导入后的项目名称"
    assert exported["项目基本情况"].cell(row=2, column=3).value == "网络安全法、数据安全法"
    assert exported["被评估单位基本信息"].cell(row=2, column=1).value == "被评估单位"
    assert exported["联系人信息"].cell(row=2, column=1).value == "张三"


def test_plan_team_excel_import_export_and_preserves_focus_gap_items(client):
    project_id = create_project(client, project_code="ENST-TEST-PLAN-TEAM-EXCEL")
    unwrap(
        client.post(
            f"/api/v1/projects/{project_id}/plan/assessment-team",
            json={"name": "旧成员", "organization": "旧单位", "role": "旧角色"},
        )
    )
    focus = unwrap(
        client.post(
            f"/api/v1/projects/{project_id}/plan/focus-points",
            json={"name": "重点关注", "domain": "制度", "riskLevel": "HIGH"},
        )
    )
    gap = unwrap(
        client.post(
            f"/api/v1/projects/{project_id}/plan/gap-items",
            json={"gapItem": "差距项", "dimension": "管理", "currentYearScore": 80},
        )
    )

    template_response = client.get(f"/api/v1/projects/{project_id}/plan/team-export-template")
    assert template_response.status_code == 200
    workbook = load_workbook(BytesIO(template_response.data))
    assert workbook.sheetnames == ["评估团队", "被评估方团队"]
    assert [cell.value for cell in workbook["评估团队"][1]] == ["姓名", "单位", "角色"]
    assert [cell.value for cell in workbook["被评估方团队"][1]] == ["公司/部门", "姓名", "职位", "联系方式"]

    workbook["评估团队"].append(["李四", "评估机构", "组长"])
    workbook["被评估方团队"].append(["信息部", "王五", "经理", "13900000000"])
    imported = unwrap(
        client.post(
            f"/api/v1/projects/{project_id}/plan/team-import",
            data={"file": (workbook_stream(workbook), "plan-team.xlsx")},
            content_type="multipart/form-data",
        )
    )
    assert imported["assessmentTeam"] == [
        {"id": imported["assessmentTeam"][0]["id"], "name": "李四", "organization": "评估机构", "role": "组长"}
    ]
    assert imported["clientTeam"] == [
        {"id": imported["clientTeam"][0]["id"], "department": "信息部", "name": "王五", "position": "经理", "contact": "13900000000"}
    ]

    session = SessionLocal()
    assessment_rows = session.query(AssessmentTeamMember).filter_by(project_id=project_id, deleted=False).all()
    client_rows = session.query(ClientTeamMember).filter_by(project_id=project_id, deleted=False).all()
    focus_rows = session.query(FocusPoint).filter_by(project_id=project_id, deleted=False).all()
    gap_rows = session.query(GapItem).filter_by(project_id=project_id, deleted=False).all()
    assert [row.name for row in assessment_rows] == ["李四"]
    assert [row.name for row in client_rows] == ["王五"]
    assert [row.id for row in focus_rows] == [focus["id"]]
    assert [row.id for row in gap_rows] == [gap["id"]]

    export_response = client.get(f"/api/v1/projects/{project_id}/plan/team-export")
    assert export_response.status_code == 200
    exported = load_workbook(BytesIO(export_response.data))
    assert exported.sheetnames == ["评估团队", "被评估方团队"]
    assert exported["评估团队"].cell(row=2, column=1).value == "李四"
    assert exported["被评估方团队"].cell(row=2, column=2).value == "王五"


def test_plan_team_import_reports_header_validation_errors(client):
    project_id = create_project(client, project_code="ENST-TEST-PLAN-TEAM-ERROR")
    workbook = Workbook()
    workbook.active.title = "评估团队"
    workbook["评估团队"].append(["姓名", "单位", "错误列"])
    client_sheet = workbook.create_sheet("被评估方团队")
    client_sheet.append(["公司/部门", "姓名", "职位", "联系方式"])

    response = client.post(
        f"/api/v1/projects/{project_id}/plan/team-import",
        data={"file": (workbook_stream(workbook), "plan-team.xlsx")},
        content_type="multipart/form-data",
    )
    assert response.status_code == 400
    payload = response.get_json()
    assert payload["code"] == "IMPORT_VALIDATION_FAILED"
    assert payload["message"] == "导入文件存在格式错误"
    assert payload["data"] == {
        "errors": [
            {
                "sheetName": "评估团队",
                "rowNo": 1,
                "field": "角色",
                "reason": "表头名称与模板不一致，请使用最新模板。",
            }
        ]
    }


def test_survey_docx_import_overwrites_lists_and_export_fills_template(client):
    project_id = create_project(client, project_code="ENST-TEST-SURVEY-DOCX")
    unwrap(
        client.post(
            f"/api/v1/projects/{project_id}/survey/business-systems",
            json={"systemName": "旧系统", "businessFunction": "旧业务"},
        )
    )
    unwrap(client.post(f"/api/v1/projects/{project_id}/survey/data-assets", json={"dataName": "旧数据资产"}))
    unwrap(client.post(f"/api/v1/projects/{project_id}/survey/personal-info", json={"dataName": "旧个人信息"}))
    unwrap(client.post(f"/api/v1/projects/{project_id}/survey/important-data", json={"dataName": "旧重要数据"}))
    unwrap(client.post(f"/api/v1/projects/{project_id}/survey/core-data", json={"dataName": "旧核心数据"}))

    survey_docx = survey_docx_stream(
        [
            (0, 1, 3, "星河电力有限公司"),
            (0, 2, 3, "91330000000000000X"),
            (0, 9, 2, "企业"),
            (0, 12, 3, "华东区域"),
            (0, 15, 3, "覆盖500万用户"),
            (1, 1, 2, "营销业务系统"),
            (1, 2, 2, "办理客户营销业务"),
            (1, 3, 2, "居民客户"),
            (1, 4, 2, "500万用户"),
            (1, 8, 2, "☑一般数据 ☑个人信息 □重要数据 ☑核心数据"),
            (2, 2, 0, "客户主数据"),
            (2, 2, 1, "数据库"),
            (2, 2, 2, "客户基础信息"),
            (2, 2, 3, "10万条"),
            (2, 2, 4, "CRM系统"),
            (2, 2, 5, "本地数据中心"),
            (2, 2, 6, "CRM流转至营销平台"),
            (2, 2, 7, "是"),
            (2, 2, 8, "个人信息"),
            (2, 2, 9, "L3"),
            (2, 2, 10, "是"),
            (3, 2, 0, "手机号"),
            (3, 2, 1, "联系方式"),
            (3, 2, 2, "10万条"),
            (3, 2, 3, "敏感"),
            (3, 2, 4, "用户注册"),
            (3, 2, 5, "注册后进入CRM"),
            (4, 2, 0, "电力交易数据"),
            (4, 2, 1, "交易类"),
            (4, 2, 2, "1万条"),
            (4, 2, 3, "交易平台"),
            (4, 2, 4, "同步至数据仓库"),
            (5, 2, 0, "调度核心数据"),
            (5, 2, 1, "调度类"),
            (5, 2, 2, "5000条"),
            (5, 2, 3, "调度系统"),
            (5, 2, 4, "仅在调度域流转"),
            (6, 1, 2, "官网、APP"),
            (6, 2, 2, "接口采集"),
            (6, 17, 2, "HTTPS"),
            (7, 1, 1, "已完成等保二级测评"),
            (7, 4, 1, "密码和多因素认证"),
        ]
    )

    imported = unwrap(
        client.post(
            f"/api/v1/projects/{project_id}/survey/import",
            data={"file": (survey_docx, "survey.docx")},
            content_type="multipart/form-data",
        )
    )
    assert imported["dataProcessorBasic"]["unitName"] == "星河电力有限公司"
    assert imported["businessSystem"]["systemName"] == "营销业务系统"
    assert imported["counts"] == {"dataAssets": 1, "personalInfo": 1, "importantData": 1, "coreData": 1}

    processor = unwrap(client.get(f"/api/v1/projects/{project_id}/survey/data-processor-basic"))
    assert processor["unitName"] == "星河电力有限公司"
    assert processor["businessScale"] == "覆盖500万用户"

    systems = unwrap(client.get(f"/api/v1/projects/{project_id}/survey/business-systems?pageNo=1&pageSize=10"))
    assert systems["total"] == 1
    assert systems["list"][0]["systemName"] == "营销业务系统"
    assert systems["list"][0]["dataScopes"] == ["一般数据", "个人信息", "核心数据"]

    assets = unwrap(client.get(f"/api/v1/projects/{project_id}/survey/data-assets?pageNo=1&pageSize=10"))
    assert assets["total"] == 1
    assert assets["list"][0]["dataName"] == "客户主数据"
    assert assets["list"][0]["dataForm"] == "数据库"
    assert assets["list"][0]["personalInfo"] is True

    personal = unwrap(client.get(f"/api/v1/projects/{project_id}/survey/personal-info?pageNo=1&pageSize=10"))
    assert personal["total"] == 1
    assert personal["list"][0]["dataName"] == "手机号"
    assert personal["list"][0]["sensitivity"] == "敏感"

    important = unwrap(client.get(f"/api/v1/projects/{project_id}/survey/important-data?pageNo=1&pageSize=10"))
    assert important["total"] == 1
    assert important["list"][0]["dataName"] == "电力交易数据"

    core = unwrap(client.get(f"/api/v1/projects/{project_id}/survey/core-data?pageNo=1&pageSize=10"))
    assert core["total"] == 1
    assert core["list"][0]["dataName"] == "调度核心数据"

    processing = unwrap(client.get(f"/api/v1/projects/{project_id}/survey/processing-activity-survey"))
    assert processing["involvedActivities"] == ["COLLECT", "TRANSFER"]
    assert processing["collectionChannels"] == "官网、APP"
    assert processing["collectionMethod"] == "接口采集"
    assert processing["onlineChannel"] == "HTTPS"

    security = unwrap(client.get(f"/api/v1/projects/{project_id}/survey/security-protection"))
    assert security["complianceAssessmentStatus"] == "已完成等保二级测评"
    assert security["identityAuthenticationAndAccessControl"] == "密码和多因素认证"

    template_response = client.get(f"/api/v1/projects/{project_id}/survey/export-template")
    assert template_response.status_code == 200
    assert template_response.mimetype == "application/vnd.openxmlformats-officedocument.wordprocessingml.document"

    export_response = client.get(f"/api/v1/projects/{project_id}/survey/export")
    assert export_response.status_code == 200
    assert export_response.mimetype == "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    with zipfile.ZipFile(BytesIO(export_response.data)) as exported:
        document_xml = exported.read("word/document.xml").decode("utf-8")
        assert "星河电力有限公司" in document_xml
        assert "营销业务系统" in document_xml
        assert "客户主数据" in document_xml
        assert "手机号" in document_xml
        assert "电力交易数据" in document_xml
        assert "调度核心数据" in document_xml
        assert "官网、APP" in document_xml
        assert "密码和多因素认证" in document_xml


def test_survey_docx_import_blank_questionnaires_clear_existing_values(client):
    project_id = create_project(client, project_code="ENST-TEST-SURVEY-DOCX-BLANK")
    unwrap(
        client.put(
            f"/api/v1/projects/{project_id}/survey/processing-activity-survey",
            json={
                "involvedActivities": ["COLLECT", "TRANSFER"],
                "collectionChannels": "旧采集渠道",
                "transferProtocol": "旧传输协议",
            },
        )
    )
    unwrap(
        client.put(
            f"/api/v1/projects/{project_id}/survey/security-protection",
            json={
                "isPowerMonitoringSystem": "YES",
                "productionControlAreaProtection": "旧生产控制区防护",
                "identityAuthenticationAndAccessControl": "旧身份鉴别",
            },
        )
    )

    unwrap(
        client.post(
            f"/api/v1/projects/{project_id}/survey/import",
            data={"file": (survey_docx_stream([]), "blank-survey.docx")},
            content_type="multipart/form-data",
        )
    )

    processing = unwrap(client.get(f"/api/v1/projects/{project_id}/survey/processing-activity-survey"))
    assert processing["involvedActivities"] == []
    assert processing["collectionChannels"] == ""
    assert processing["transferProtocol"] == ""

    security = unwrap(client.get(f"/api/v1/projects/{project_id}/survey/security-protection"))
    assert security["isPowerMonitoringSystem"] == "NO"
    assert security["productionControlAreaProtection"] == ""
    assert security["identityAuthenticationAndAccessControl"] == ""


def test_survey_docx_import_saves_security_answers_after_stripping_template_text(client):
    project_id = create_project(client, project_code="ENST-TEST-SURVEY-DOCX-SECURITY-SAME")

    unwrap(
        client.post(
            f"/api/v1/projects/{project_id}/survey/import",
            data={
                "file": (
                    survey_docx_stream(
                        [
                            (
                                7,
                                1,
                                1,
                                "已开展的等级保护测评、商用密码应用安全性评估、安全检测、风险评估、安全认证、合规审计情况，及发现问题的整改情况。"
                                "已开展了等级保护测评、上线前安全检测、功能检测、源代码检测、漏洞扫描。",
                            ),
                            (
                                7,
                                4,
                                1,
                                "身份鉴别与访问控制情况"
                                "应用层面，用户通过账号和密码进行登录，已配置用户口令长度不小于9位，大小写字母、数字、特殊字符组成，用户名与口令不同，口令加密存储，口令每180天更换一次。"
                                "数据库层面，通过账号和密码直接登录，配置了口令复杂度要求，最少9位，必须由大写字母、数字、特殊字符组成，并要求定期更换。",
                            ),
                            (7, 5, 1, "网络安全漏洞管理及修复情况针对漏洞扫描中发现的高危漏洞进行了整改。"),
                            (7, 6, 1, "VPN等远程管理软件的用户及管理情况不涉及"),
                            (
                                7,
                                8,
                                1,
                                "加密、脱敏、去标识化等安全技术应用情况"
                                "本系统不涉及敏感个人信息和重要数据，对用户表中的口令进行了加密处理。前端展示中，在系统管理-用户管理中，对展示的用户列表中的手机号进行了脱敏；"
                                "在数据看板和驾驶舱中，因需实时展示值班人员信息，未值班人员的手机号进行脱敏。",
                            ),
                        ]
                    ),
                    "security-same-as-template.docx",
                )
            },
            content_type="multipart/form-data",
        )
    )

    security = unwrap(client.get(f"/api/v1/projects/{project_id}/survey/security-protection"))
    assert security["complianceAssessmentStatus"] == "已开展了等级保护测评、上线前安全检测、功能检测、源代码检测、漏洞扫描。"
    assert security["identityAuthenticationAndAccessControl"].startswith("应用层面，用户通过账号和密码进行登录")
    assert security["vulnerabilityManagement"] == "针对漏洞扫描中发现的高危漏洞进行了整改。"
    assert security["remoteManagementSoftware"] == "不涉及"
    assert security["securityTechnologyApplication"].startswith("本系统不涉及敏感个人信息和重要数据")
    assert security["isPowerMonitoringSystem"] == "NO"

    session = SessionLocal()
    row = session.query(SecurityProtectionSurvey).filter_by(project_id=project_id, deleted=False).one()
    assert row.compliance_assessment_status == security["complianceAssessmentStatus"]
    assert row.identity_authentication_and_access_control == security["identityAuthenticationAndAccessControl"]


def test_survey_docx_import_none_markers_clear_asset_tables(client):
    project_id = create_project(client, project_code="ENST-TEST-SURVEY-DOCX-ASSET-NONE")
    unwrap(client.post(f"/api/v1/projects/{project_id}/survey/data-assets", json={"dataName": "旧数据资产"}))
    unwrap(client.post(f"/api/v1/projects/{project_id}/survey/personal-info", json={"dataName": "旧个人信息"}))
    unwrap(client.post(f"/api/v1/projects/{project_id}/survey/important-data", json={"dataName": "旧重要数据"}))
    unwrap(client.post(f"/api/v1/projects/{project_id}/survey/core-data", json={"dataName": "旧核心数据"}))

    imported = unwrap(
        client.post(
            f"/api/v1/projects/{project_id}/survey/import",
            data={
                "file": (
                    survey_docx_stream(
                        [
                            (2, 2, 0, "（ 不 涉 及 。）"),
                            (3, 2, 0, "【无】"),
                            (4, 2, 0, " 无 "),
                            (5, 2, 0, "不 涉 及"),
                        ]
                    ),
                    "asset-none-survey.docx",
                )
            },
            content_type="multipart/form-data",
        )
    )

    assert imported["counts"] == {"dataAssets": 0, "personalInfo": 0, "importantData": 0, "coreData": 0}
    assert unwrap(client.get(f"/api/v1/projects/{project_id}/survey/data-assets?pageNo=1&pageSize=10"))["total"] == 0
    assert unwrap(client.get(f"/api/v1/projects/{project_id}/survey/personal-info?pageNo=1&pageSize=10"))["total"] == 0
    assert unwrap(client.get(f"/api/v1/projects/{project_id}/survey/important-data?pageNo=1&pageSize=10"))["total"] == 0
    assert unwrap(client.get(f"/api/v1/projects/{project_id}/survey/core-data?pageNo=1&pageSize=10"))["total"] == 0


def test_survey_docx_import_ignores_none_markers_when_deriving_processing_activities(client):
    project_id = create_project(client, project_code="ENST-TEST-SURVEY-DOCX-PROCESSING-NONE")

    unwrap(
        client.post(
            f"/api/v1/projects/{project_id}/survey/import",
            data={
                "file": (
                    survey_docx_stream(
                        [
                            (6, 1, 2, "（ 不 涉 及 。）"),
                            (6, 2, 2, "【无】"),
                            (6, 9, 2, " □ 无 "),
                            (6, 17, 2, "HTTPS"),
                        ]
                    ),
                    "processing-none-survey.docx",
                )
            },
            content_type="multipart/form-data",
        )
    )

    processing = unwrap(client.get(f"/api/v1/projects/{project_id}/survey/processing-activity-survey"))
    assert processing["involvedActivities"] == ["TRANSFER"]
    assert processing["collectionChannels"] == "（ 不 涉 及 。）"
    assert processing["collectionMethod"] == "【无】"
    assert processing["collectionPublicDeviceUsage"] == "□ 无"
    assert processing["onlineChannel"] == "HTTPS"


def test_survey_docx_import_reads_decomposed_power_monitoring_protection_rows(client):
    project_id = create_project(client, project_code="ENST-TEST-SURVEY-DOCX-POWER-DECOMPOSED")

    unwrap(
        client.post(
            f"/api/v1/projects/{project_id}/survey/import",
            data={
                "file": (
                    survey_docx_power_decomposed_stream(
                        [
                            ["9", "是否为电力监控系统", "是"],
                            ["", "生产控制区和管理信息区的设置和防护情况", "生产区已分区防护"],
                            ["", "安全接入区的设立情况", "已设安全接入区"],
                            ["", "电力监控专用网络的使用情况", "使用专用网络"],
                            ["", "生产控制区与管理信息区、安全接入区的隔离及隔离装置使用情况", "部署隔离装置"],
                            ["", "生产控制区与电力监控专用网络的广域网之间的联接安全方案", "广域网双向认证"],
                            ["", "电力调度认证机制建设情况", "已部署调度证书"],
                            ["", "网络服务的安全管控情况", "最小化开放服务"],
                            ["", "安全接入区的安全管控情况", "接入区双因素认证"],
                            ["", "电力监控系统分区边界的安全防护情况", "边界部署访问控制"],
                            ["", "系统使用的产品安全可靠情况", "使用可信产品"],
                            ["", "运营者的网络安全检测预警机制建设情况", "7x24 检测预警"],
                        ],
                        threat_value="3年内未发生重大事件",
                    ),
                    "power-decomposed-survey.docx",
                )
            },
            content_type="multipart/form-data",
        )
    )

    security = unwrap(client.get(f"/api/v1/projects/{project_id}/survey/security-protection"))
    assert security["isPowerMonitoringSystem"] == "YES"
    assert security["productionControlAreaProtection"] == "生产区已分区防护"
    assert security["securityAccessAreaSetup"] == "已设安全接入区"
    assert security["powerMonitoringDedicatedNetwork"] == "使用专用网络"
    assert security["zoneIsolationDeviceUsage"] == "部署隔离装置"
    assert security["wideAreaNetworkConnectionSecurity"] == "广域网双向认证"
    assert security["powerDispatchAuthentication"] == "已部署调度证书"
    assert security["networkServiceSecurityControl"] == "最小化开放服务"
    assert security["securityAccessAreaSecurityControl"] == "接入区双因素认证"
    assert security["zoneBoundaryProtection"] == "边界部署访问控制"
    assert security["productSecurityReliability"] == "使用可信产品"
    assert security["operatorSecurityMonitoringWarning"] == "7x24 检测预警"
    assert security["securityIncidentsAndThreats"] == "3年内未发生重大事件"


def test_survey_docx_import_power_monitoring_rows_strip_template_text_and_fallback_by_position(client):
    project_id = create_project(client, project_code="ENST-TEST-SURVEY-DOCX-POWER-POSITION")

    unwrap(
        client.post(
            f"/api/v1/projects/{project_id}/survey/import",
            data={
                "file": (
                    survey_docx_power_decomposed_stream(
                        [
                            ["9", "是否为电力监控系统： ☑是  □否"],
                            ["", "1）生产控制区和管理信息区的设置和防护情况 是积分卡flask就"],
                            ["", "安全接入区的设立情况艰苦拉萨酱豆腐"],
                            ["", "啊实打实大苏打"],
                            ["", "4）生产控制区与管理信息区、安全接入区的隔离及隔离装置使用情况"],
                            ["", "5）生产控制区与电力监控专用网络的广域网之间的连接安全方案"],
                            ["", "6）电力调度认证机制建设情况"],
                            ["", "7）网络服务的安全管控情况8）安全接入区的安全管控情况"],
                            ["", "9）电力监控系统分区边界的安全防护情况"],
                            ["", "10）系统使用的产品安全可靠情况"],
                            ["", "11）运营者的网络安全监测预警机制建设情况"],
                        ]
                    ),
                    "power-position-survey.docx",
                )
            },
            content_type="multipart/form-data",
        )
    )

    security = unwrap(client.get(f"/api/v1/projects/{project_id}/survey/security-protection"))
    assert security["isPowerMonitoringSystem"] == "YES"
    assert security["productionControlAreaProtection"] == "是积分卡flask就"
    assert security["securityAccessAreaSetup"] == "艰苦拉萨酱豆腐"
    assert security["powerMonitoringDedicatedNetwork"] == "啊实打实大苏打"
    assert security["networkServiceSecurityControl"] == ""
    assert security["securityAccessAreaSecurityControl"] == ""
    assert security["zoneBoundaryProtection"] == ""


def test_survey_docx_import_skips_power_monitoring_detail_rows_when_marked_no(client):
    project_id = create_project(client, project_code="ENST-TEST-SURVEY-DOCX-POWER-NO")
    unwrap(
        client.put(
            f"/api/v1/projects/{project_id}/survey/security-protection",
            json={
                "isPowerMonitoringSystem": "YES",
                "productionControlAreaProtection": "旧生产控制区防护",
                "securityAccessAreaSetup": "旧安全接入区",
                "operatorSecurityMonitoringWarning": "旧检测预警",
            },
        )
    )

    unwrap(
        client.post(
            f"/api/v1/projects/{project_id}/survey/import",
            data={
                "file": (
                    survey_docx_power_decomposed_stream(
                        [
                            ["9", "是否为电力监控系统： □是  ☑否"],
                            ["", "1）生产控制区和管理信息区的设置和防护情况", "不应导入生产区"],
                            ["", "2）安全接入区的设立情况", "不应导入接入区"],
                            ["", "3）电力监控专用网络的使用情况", "不应导入专网"],
                            ["", "4）生产控制区与管理信息区、安全接入区的隔离及隔离装置使用情况", "不应导入隔离"],
                            ["", "5）生产控制区与电力监控专用网络的广域网之间的连接安全方案", "不应导入广域网"],
                            ["", "6）电力调度认证机制建设情况", "不应导入调度认证"],
                            ["", "7）网络服务的安全管控情况", "不应导入网络服务"],
                            ["", "8）安全接入区的安全管控情况", "不应导入安全接入区管控"],
                            ["", "9）电力监控系统分区边界的安全防护情况", "不应导入边界"],
                            ["", "10）系统使用的产品安全可靠情况", "不应导入产品"],
                            ["", "11）运营者的网络安全监测预警机制建设情况", "不应导入预警"],
                        ],
                        threat_value="3年内未发生安全事件",
                    ),
                    "power-no-survey.docx",
                )
            },
            content_type="multipart/form-data",
        )
    )

    security = unwrap(client.get(f"/api/v1/projects/{project_id}/survey/security-protection"))
    assert security["isPowerMonitoringSystem"] == "NO"
    assert security["productionControlAreaProtection"] == ""
    assert security["securityAccessAreaSetup"] == ""
    assert security["powerMonitoringDedicatedNetwork"] == ""
    assert security["zoneIsolationDeviceUsage"] == ""
    assert security["wideAreaNetworkConnectionSecurity"] == ""
    assert security["powerDispatchAuthentication"] == ""
    assert security["networkServiceSecurityControl"] == ""
    assert security["securityAccessAreaSecurityControl"] == ""
    assert security["zoneBoundaryProtection"] == ""
    assert security["productSecurityReliability"] == ""
    assert security["operatorSecurityMonitoringWarning"] == ""
    assert security["securityIncidentsAndThreats"] == "3年内未发生安全事件"


def test_survey_docx_import_rejects_non_docx(client):
    project_id = create_project(client, project_code="ENST-TEST-SURVEY-DOCX-INVALID")

    response = client.post(
        f"/api/v1/projects/{project_id}/survey/import",
        data={"file": (BytesIO(b"not a word file"), "survey.txt")},
        content_type="multipart/form-data",
    )

    assert response.status_code == 400
    payload = response.get_json()
    assert payload["code"] == "IMPORT_VALIDATION_FAILED"
    assert payload["data"]["errors"] == [
        {
            "tableName": None,
            "rowNo": 0,
            "field": "file",
            "reason": "导入文件必须是 .docx 格式。",
        }
    ]


def test_business_system_diagram_upload_updates_file_ids(client):
    project_id = create_project(client, project_code="ENST-TEST-DIAGRAM")
    business_system = unwrap(
        client.post(
            f"/api/v1/projects/{project_id}/survey/business-systems",
            json={"systemName": "Customer Platform", "businessFunction": "Customer management"},
        )
    )
    record_id = business_system["id"]

    topology_bytes = b"\x89PNG\r\n\x1a\ntopology"
    topology = unwrap(
        client.post(
            f"/api/v1/projects/{project_id}/survey/business-systems/{record_id}/topology-diagram",
            data={"file": (BytesIO(topology_bytes), "topology.png")},
            content_type="multipart/form-data",
        )
    )
    assert topology["file"]["bizType"] == "SURVEY_TOPOLOGY_DIAGRAM"
    assert topology["file"]["storageProvider"] == "LOCAL"
    assert topology["businessSystem"]["topologyFileId"] == topology["file"]["fileId"]

    data_flow_bytes = b"\x89PNG\r\n\x1a\ndata-flow"
    data_flow = unwrap(
        client.post(
            f"/api/v1/projects/{project_id}/survey/business-systems/{record_id}/data-flow-diagram",
            data={"file": (BytesIO(data_flow_bytes), "data-flow.png")},
            content_type="multipart/form-data",
        )
    )
    assert data_flow["file"]["bizType"] == "SURVEY_DATA_FLOW_DIAGRAM"
    assert data_flow["businessSystem"]["businessFlowFileId"] == data_flow["file"]["fileId"]

    vsdx_bytes = b"PK\x03\x04vsdx"
    vsdx = unwrap(
        client.post(
            f"/api/v1/projects/{project_id}/survey/business-systems/{record_id}/topology-diagram",
            data={"file": (BytesIO(vsdx_bytes), "topology.vsdx")},
            content_type="multipart/form-data",
        )
    )
    assert vsdx["file"]["fileName"] == "topology.vsdx"
    assert vsdx["businessSystem"]["topologyFileId"] == vsdx["file"]["fileId"]

    stored = unwrap(client.get(f"/api/v1/projects/{project_id}/survey/business-systems?pageNo=1&pageSize=10"))
    assert stored["list"][0]["topologyFileId"] == vsdx["file"]["fileId"]
    assert stored["list"][0]["businessFlowFileId"] == data_flow["file"]["fileId"]

    download = client.get(topology["file"]["downloadUrl"])
    assert download.status_code == 200
    assert download.data == topology_bytes

    invalid = client.post(
        f"/api/v1/projects/{project_id}/survey/business-systems/{record_id}/data-flow-diagram",
        data={"file": (BytesIO(b"not image"), "data-flow.txt")},
        content_type="multipart/form-data",
    )
    assert invalid.status_code == 400
    assert invalid.get_json()["code"] == "INVALID_DIAGRAM_FILE"


def test_harm_analysis_uses_dashscope_gateway_before_fallback(client, monkeypatch):
    from app.services import llm_gateway_service

    project_id = create_project(client, project_code="ENST-TEST-LLM")
    unwrap(client.post(f"/api/v1/projects/{project_id}/start"))
    items = unwrap(client.get(f"/api/v1/projects/{project_id}/evaluation/items?pageNo=1&pageSize=1"))
    item_id = items["list"][0]["id"]
    unwrap(
        client.put(
            f"/api/v1/projects/{project_id}/evaluation/items/{item_id}/record",
            json={"evaluationResult": "PARTIAL", "evaluationRecord": "record"},
        )
    )
    unwrap(client.post(f"/api/v1/projects/{project_id}/risk-summary/refresh", json={}))
    risk_items = unwrap(client.get(f"/api/v1/projects/{project_id}/risk-items?pageNo=1&pageSize=1"))
    risk_item_id = risk_items["list"][0]["id"]

    def fake_suggest_harm_analysis(_project, _record, _system_category, _rule_config):
        assert _record.assessment_category
        assert _record.assessment_subcategory
        assert _record.check_point
        assert _record.evaluation_result == "PARTIAL"
        assert _record.evaluation_record == "record"
        return {
            "impactedObject": "PUBLIC_INTEREST",
            "damageDegree": "SERIOUS",
            "reason": "模型判断该风险会影响公共利益。",
            "confidence": 0.91,
            "needsManualReview": False,
            "llm_status": llm_gateway_service.LLM_SUCCESS,
            "llm_provider": "dashscope",
            "llm_model": "qwen-plus",
        }

    monkeypatch.setattr(llm_gateway_service, "suggest_harm_analysis", fake_suggest_harm_analysis)
    suggestion = unwrap(client.post(f"/api/v1/projects/{project_id}/risk-items/{risk_item_id}/harm-analysis/suggest", json={}))
    assert suggestion["llmStatus"] == "SUCCESS"
    assert suggestion["llmProvider"] == "dashscope"
    assert suggestion["llmModel"] == "qwen-plus"
    assert suggestion["harmLevel"] == "RELATIVELY_HIGH"

    applied = unwrap(
        client.post(
            f"/api/v1/projects/{project_id}/risk-items/{risk_item_id}/harm-analysis/apply",
            json={"suggestion": suggestion},
        )
    )
    assert applied["harmLevel"] == "RELATIVELY_HIGH"
    assert applied["riskLevel"] == "MEDIUM"


def test_harm_analysis_prompt_includes_evaluation_context(client, monkeypatch):
    from app.services import llm_gateway_service

    project_id = create_project(client, project_code="ENST-TEST-PROMPT")
    unwrap(client.post(f"/api/v1/projects/{project_id}/start"))
    items = unwrap(client.get(f"/api/v1/projects/{project_id}/evaluation/items?pageNo=1&pageSize=1"))
    item = items["list"][0]
    unwrap(
        client.put(
            f"/api/v1/projects/{project_id}/evaluation/items/{item['id']}/record",
            json={"evaluationResult": "NON_COMPLIANT", "evaluationRecord": "现场记录用于区分严重性"},
        )
    )
    unwrap(client.post(f"/api/v1/projects/{project_id}/risk-summary/refresh", json={}))
    risk_items = unwrap(client.get(f"/api/v1/projects/{project_id}/risk-items?pageNo=1&pageSize=1"))
    risk_item_id = risk_items["list"][0]["id"]
    unwrap(
        client.put(
            f"/api/v1/projects/{project_id}/risk-sources/{risk_item_id}",
            json={
                "relatedData": ["用户身份信息/L3", "业务交易数据/L2"],
                "relatedActivities": ["收集", "存储", "使用和加工"],
            },
        )
    )
    captured = {}

    def fake_chat_completion(_config, messages):
        import json

        prompt = json.loads(messages[1]["content"])
        captured.update(prompt["input"])
        return json.dumps(
            {
                "impactedObject": "LEGAL_RIGHTS",
                "damageDegree": "SERIOUS",
                "reason": "根据现场测评上下文判断。",
                "harmAnalysis": "现场未落实访问控制，导致用户身份信息在存储和使用环节存在未授权访问风险，可能严重侵害用户合法权益，因此危害程度为中。",
                "confidence": 0.8,
                "needsManualReview": False,
            },
            ensure_ascii=False,
        )

    monkeypatch.setattr(llm_gateway_service, "_chat_completion", fake_chat_completion)
    suggestion = unwrap(client.post(f"/api/v1/projects/{project_id}/risk-items/{risk_item_id}/harm-analysis/suggest", json={}))
    assert suggestion["llmStatus"] == "SUCCESS"
    assert captured["assessmentCategory"] == item["category"]
    assert captured["assessmentSubcategory"] == item["subcategory"]
    assert captured["checkPoint"] == item["checkPoint"]
    assert captured["evaluationResult"] == "NON_COMPLIANT"
    assert captured["evaluationRecord"] == "现场记录用于区分严重性"
    assert captured["relatedData"] == "用户身份信息/L3、业务交易数据/L2"
    assert captured["relatedActivities"] == ["收集", "存储", "使用和加工"]
    assert suggestion["harmDescription"] == "现场未落实访问控制，导致用户身份信息在存储和使用环节存在未授权访问风险，可能严重侵害用户合法权益，因此危害程度为中。"
    assert "涉及的数据及类型、级别：用户身份信息/L3、业务交易数据/L2" in suggestion["harmAnalysisTrace"]["step2Basis"]
    assert "涉及的数据处理活动：收集、存储、使用和加工" in suggestion["harmAnalysisTrace"]["step2Basis"]

    unwrap(
        client.put(
            f"/api/v1/projects/{project_id}/risk-sources/{risk_item_id}",
            json={"relatedData": ["用户身份信息/L4"]},
        )
    )
    updated_suggestion = unwrap(
        client.post(f"/api/v1/projects/{project_id}/risk-items/{risk_item_id}/harm-analysis/suggest", json={})
    )
    assert updated_suggestion["harmAnalysisInputHash"] != suggestion["harmAnalysisInputHash"]


def test_evaluation_import_export_and_report_management(client):
    project_id = create_project(client)
    unwrap(client.post(f"/api/v1/projects/{project_id}/start"))

    template_response = client.get(f"/api/v1/projects/{project_id}/evaluation/export-template")
    assert template_response.status_code == 200
    assert template_response.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    workbook = load_workbook(BytesIO(template_response.data))
    worksheet = workbook.active
    assert worksheet.cell(row=1, column=3).value == "评估项ID"
    assert worksheet.cell(row=1, column=8).value == "评估结果"
    assert worksheet.cell(row=1, column=9).value == "符合情况"
    assert worksheet.cell(row=2, column=1).value
    assert worksheet.cell(row=2, column=3).value == "AQGL001"
    worksheet.cell(row=2, column=8).value = "imported record"
    worksheet.cell(row=2, column=9).value = "PARTIAL"
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    imported = unwrap(
        client.post(
            f"/api/v1/projects/{project_id}/evaluation/import",
            data={"file": (stream, "evaluation-import.xlsx")},
            content_type="multipart/form-data",
        )
    )
    assert imported["importedCount"] == 1
    assert imported["failedCount"] == 0
    session = SessionLocal()
    records = (
        session.query(EvaluationRecord)
        .filter(EvaluationRecord.project_id == project_id, EvaluationRecord.deleted.is_(False))
        .all()
    )
    assert len(records) == 1
    record_id = records[0].id

    worksheet.cell(row=2, column=8).value = "updated record"
    worksheet.cell(row=2, column=9).value = "NON_COMPLIANT"
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    reimported = unwrap(
        client.post(
            f"/api/v1/projects/{project_id}/evaluation/import",
            data={"file": (stream, "evaluation-import.xlsx")},
            content_type="multipart/form-data",
        )
    )
    assert reimported["importedCount"] == 1
    records = (
        session.query(EvaluationRecord)
        .filter(EvaluationRecord.project_id == project_id, EvaluationRecord.deleted.is_(False))
        .all()
    )
    assert len(records) == 1
    assert records[0].id == record_id
    assert records[0].evaluation_record == "updated record"
    assert records[0].evaluation_result == "NON_COMPLIANT"

    export_response = client.get(f"/api/v1/projects/{project_id}/evaluation/export")
    assert export_response.status_code == 200
    exported = load_workbook(BytesIO(export_response.data)).active
    assert exported.max_column == 9
    assert exported.cell(row=2, column=3).value == "AQGL001"
    assert exported.cell(row=2, column=8).value == "updated record"
    assert exported.cell(row=2, column=9).value == "NON_COMPLIANT"
    unwrap(
        client.post(
            f"/api/v1/projects/{project_id}/survey/business-systems",
            json={"systemName": "Flow Project", "businessFunction": "Flow report business"},
        )
    )

    generated = unwrap(
        client.post(
            f"/api/v1/projects/{project_id}/reports/generate",
            json={"reportName": "Flow Report", "selectedSections": ["EVALUATION", "RISK_SUMMARY"]},
        )
    )
    assert generated["status"] == "SUCCESS"
    assert generated["reportId"]

    task = unwrap(client.get(f"/api/v1/projects/{project_id}/reports/tasks/{generated['reportTaskId']}"))
    assert task["status"] == "SUCCESS"

    reports = unwrap(client.get(f"/api/v1/projects/{project_id}/reports?pageNo=1&pageSize=10"))
    assert reports["total"] == 1
    assert reports["list"][0]["downloadUrl"]

    download = client.get(f"/api/v1/projects/{project_id}/reports/{generated['reportId']}/download")
    assert download.status_code == 200
    assert download.mimetype == "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    with zipfile.ZipFile(BytesIO(download.data)) as report:
        document_xml = report.read("word/document.xml").decode("utf-8")
        assert "Flow Project" in document_xml
        document = ET.fromstring(document_xml)
        paragraph_texts = [
            "".join(node.text or "" for node in paragraph.iter("{http://schemas.openxmlformats.org/wordprocessingml/2006/main}t")).strip()
            for paragraph in document.iter("{http://schemas.openxmlformats.org/wordprocessingml/2006/main}p")
        ]
        assert "数据安全基本信息" not in paragraph_texts
        assert "安全风险处理" not in paragraph_texts
        assert not any(name.startswith("word/comments") for name in report.namelist())

    deleted = unwrap(client.delete(f"/api/v1/projects/{project_id}/reports/{generated['reportId']}"))
    assert deleted["reportId"] == generated["reportId"]
    assert unwrap(client.get(f"/api/v1/projects/{project_id}/reports?pageNo=1&pageSize=10"))["total"] == 0


def test_evaluation_import_resolves_blank_item_id_by_context(client):
    project_id = create_project(client)
    unwrap(client.post(f"/api/v1/projects/{project_id}/start"))

    template_response = client.get(f"/api/v1/projects/{project_id}/evaluation/export-template")
    workbook = load_workbook(BytesIO(template_response.data))
    worksheet = workbook.active

    rows_by_code = {}
    duplicate_rows = None
    for row_no in range(2, worksheet.max_row + 1):
        item_code = worksheet.cell(row=row_no, column=2).value
        if item_code in rows_by_code:
            duplicate_rows = (rows_by_code[item_code], row_no)
            break
        rows_by_code[item_code] = row_no
    assert duplicate_rows

    expected_item_ids = {worksheet.cell(row=row_no, column=1).value for row_no in duplicate_rows}
    for row_no in duplicate_rows:
        worksheet.cell(row=row_no, column=1).value = None
        worksheet.cell(row=row_no, column=8).value = f"legacy import row {row_no}"
        worksheet.cell(row=row_no, column=9).value = "PARTIAL"

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    imported = unwrap(
        client.post(
            f"/api/v1/projects/{project_id}/evaluation/import",
            data={"file": (stream, "evaluation-import.xlsx")},
            content_type="multipart/form-data",
        )
    )
    assert imported["importedCount"] == 2
    assert imported["failedCount"] == 0

    session = SessionLocal()
    records = (
        session.query(EvaluationRecord)
        .filter(EvaluationRecord.project_id == project_id, EvaluationRecord.deleted.is_(False))
        .all()
    )
    assert {record.item_id for record in records} == expected_item_ids
    assert {record.evaluation_record for record in records} == {f"legacy import row {row_no}" for row_no in duplicate_rows}


def test_evaluation_import_reports_invalid_result_and_preserves_records(client):
    project_id = create_project(client)
    unwrap(client.post(f"/api/v1/projects/{project_id}/start"))

    template_response = client.get(f"/api/v1/projects/{project_id}/evaluation/export-template")
    workbook = load_workbook(BytesIO(template_response.data))
    worksheet = workbook.active
    worksheet.cell(row=2, column=8).value = "original record"
    worksheet.cell(row=2, column=9).value = "PARTIAL"

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    imported = unwrap(
        client.post(
            f"/api/v1/projects/{project_id}/evaluation/import",
            data={"file": (stream, "evaluation-import.xlsx")},
            content_type="multipart/form-data",
        )
    )
    assert imported["importedCount"] == 1

    worksheet.cell(row=2, column=8).value = "should not overwrite"
    worksheet.cell(row=2, column=9).value = "错误符合情况"
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    failed = unwrap(
        client.post(
            f"/api/v1/projects/{project_id}/evaluation/import",
            data={"file": (stream, "evaluation-import.xlsx")},
            content_type="multipart/form-data",
        )
    )

    assert failed["importedCount"] == 0
    assert failed["failedCount"] == 1
    assert failed["errors"] == [
        {
            "rowNo": 2,
            "field": "符合情况",
            "reason": "符合情况列只能填写符合、基本符合、不符合、不适用或对应枚举值 COMPLIANT、PARTIAL、NON_COMPLIANT、NOT_APPLICABLE。",
        }
    ]
    session = SessionLocal()
    record = (
        session.query(EvaluationRecord)
        .filter(EvaluationRecord.project_id == project_id, EvaluationRecord.deleted.is_(False))
        .one()
    )
    assert record.evaluation_record == "original record"
    assert record.evaluation_result == "PARTIAL"


def test_evaluation_import_reports_readonly_column_changes_and_preserves_records(client):
    project_id = create_project(client, project_code="ENST-TEST-IMPORT-READONLY")
    unwrap(client.post(f"/api/v1/projects/{project_id}/start"))

    template_response = client.get(f"/api/v1/projects/{project_id}/evaluation/export-template")
    workbook = load_workbook(BytesIO(template_response.data))
    worksheet = workbook.active
    worksheet.cell(row=2, column=8).value = "original record"
    worksheet.cell(row=2, column=9).value = "PARTIAL"
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    imported = unwrap(
        client.post(
            f"/api/v1/projects/{project_id}/evaluation/import",
            data={"file": (stream, "evaluation-import.xlsx")},
            content_type="multipart/form-data",
        )
    )
    assert imported["importedCount"] == 1

    worksheet.cell(row=2, column=2).value = "edited item code"
    worksheet.cell(row=2, column=3).value = "edited assessment item id"
    worksheet.cell(row=2, column=4).value = "edited sheet"
    worksheet.cell(row=2, column=5).value = "edited category"
    worksheet.cell(row=2, column=6).value = "edited subcategory"
    worksheet.cell(row=2, column=7).value = "edited check point"
    worksheet.cell(row=2, column=8).value = "should not overwrite"
    worksheet.cell(row=2, column=9).value = "NON_COMPLIANT"
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    failed = unwrap(
        client.post(
            f"/api/v1/projects/{project_id}/evaluation/import",
            data={"file": (stream, "evaluation-import.xlsx")},
            content_type="multipart/form-data",
        )
    )

    assert failed["importedCount"] == 0
    assert failed["failedCount"] == 6
    assert failed["errors"] == [
        {"rowNo": 2, "field": "检查项编号", "reason": "检查项编号与系统检查项不一致，请使用最新导出模板。"},
        {"rowNo": 2, "field": "评估项ID", "reason": "评估项ID与系统检查项不一致，请使用最新导出模板。"},
        {"rowNo": 2, "field": "工作表", "reason": "工作表与系统检查项不一致，请使用最新导出模板。"},
        {"rowNo": 2, "field": "一级分类", "reason": "一级分类与系统检查项不一致，请使用最新导出模板。"},
        {"rowNo": 2, "field": "二级分类", "reason": "二级分类与系统检查项不一致，请使用最新导出模板。"},
        {"rowNo": 2, "field": "检查要点", "reason": "检查要点与系统检查项不一致，请使用最新导出模板。"},
    ]
    session = SessionLocal()
    record = (
        session.query(EvaluationRecord)
        .filter(EvaluationRecord.project_id == project_id, EvaluationRecord.deleted.is_(False))
        .one()
    )
    assert record.evaluation_record == "original record"
    assert record.evaluation_result == "PARTIAL"
