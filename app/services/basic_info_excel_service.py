from __future__ import annotations

import re
from datetime import date, datetime

from openpyxl import Workbook

from app.common.utils import new_id
from app.extensions import SessionLocal
from app.models import ProjectBasicInfo, ProjectReference
from app.services import basic_info_service
from app.services.excel_common import (
    load_import_workbook,
    normalize_cell,
    raise_import_validation,
    set_column_widths,
    validate_workbook_schema,
    workbook_response,
)
from app.services.project_service import get_project


BASIC_SHEET = "项目基本情况"
ORGANIZATION_SHEET = "被评估单位基本信息"
CONTACT_SHEET = "联系人信息"

BASIC_COLUMNS = [
    ("projectNumber", "项目编号"),
    ("projectName", "项目名称"),
    ("laws", "评估所依据的法律法规"),
    ("standards", "评估所参考的标准规范"),
    ("startDate", "评估开始日期"),
    ("endDate", "评估结束日期"),
]
ORGANIZATION_COLUMNS = [
    ("name", "单位名称"),
    ("postalCode", "邮政编码"),
]
CONTACT_COLUMNS = [
    ("name", "姓名"),
    ("department", "所属部门"),
    ("mobile", "移动电话"),
    ("title", "职务/职称"),
    ("phone", "办公电话"),
    ("email", "电子邮件"),
]

SHEET_SPECS = [
    (BASIC_SHEET, [label for _key, label in BASIC_COLUMNS]),
    (ORGANIZATION_SHEET, [label for _key, label in ORGANIZATION_COLUMNS]),
    (CONTACT_SHEET, [label for _key, label in CONTACT_COLUMNS]),
]

SPLIT_PATTERN = re.compile(r"[、,，;；\r\n]+")
EMAIL_PATTERN = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


def export_template_workbook(project_id: str) -> tuple[str, bytes, str]:
    get_project(project_id)
    return workbook_response(f"项目基本信息模板-{project_id}.xlsx", _build_workbook())


def export_basic_info_workbook(project_id: str) -> tuple[str, bytes, str]:
    get_project(project_id)
    data = basic_info_service.get_basic_info(project_id)
    return workbook_response(f"项目基本信息-{project_id}.xlsx", _build_workbook(data))


def import_basic_info_workbook(project_id: str, file) -> dict:
    project = get_project(project_id)
    workbook = load_import_workbook(file)
    validate_workbook_schema(workbook, SHEET_SPECS)

    errors = []
    basic_values = _row_values(workbook[BASIC_SHEET], BASIC_COLUMNS, 2)
    start_date = _parse_date(basic_values.get("startDate"), BASIC_SHEET, "评估开始日期", errors)
    end_date = _parse_date(basic_values.get("endDate"), BASIC_SHEET, "评估结束日期", errors)
    organization = _row_values(workbook[ORGANIZATION_SHEET], ORGANIZATION_COLUMNS, 2)
    contacts = _parse_contacts(workbook[CONTACT_SHEET], errors)
    if errors:
        raise_import_validation(errors)

    reference_maps = _current_reference_maps(project_id)
    return {
        "projectNumber": basic_values.get("projectNumber") or project.project_code,
        "projectName": basic_values.get("projectName") or project.project_name,
        "laws": _reference_items(basic_values.get("laws"), reference_maps["laws"]),
        "standards": _reference_items(basic_values.get("standards"), reference_maps["standards"]),
        "assessmentPlan": {"startDate": start_date, "endDate": end_date},
        "organization": organization,
        "contacts": contacts,
    }


def _build_workbook(data: dict | None = None) -> Workbook:
    workbook = Workbook()
    basic_sheet = workbook.active
    basic_sheet.title = BASIC_SHEET
    _append_headers(basic_sheet, BASIC_COLUMNS)
    organization_sheet = workbook.create_sheet(ORGANIZATION_SHEET)
    _append_headers(organization_sheet, ORGANIZATION_COLUMNS)
    contact_sheet = workbook.create_sheet(CONTACT_SHEET)
    _append_headers(contact_sheet, CONTACT_COLUMNS)

    if data:
        plan = data.get("assessmentPlan") or {}
        basic_sheet.append(
            [
                data.get("projectNumber"),
                data.get("projectName"),
                _join_names(data.get("laws") or []),
                _join_names(data.get("standards") or []),
                plan.get("startDate"),
                plan.get("endDate"),
            ]
        )
        organization = data.get("organization") or {}
        organization_sheet.append([organization.get(key) for key, _label in ORGANIZATION_COLUMNS])
        for contact in data.get("contacts") or []:
            contact_sheet.append([contact.get(key) for key, _label in CONTACT_COLUMNS])

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        set_column_widths(sheet)
    return workbook


def _append_headers(worksheet, columns: list[tuple[str, str]]) -> None:
    worksheet.append([label for _key, label in columns])


def _row_values(worksheet, columns: list[tuple[str, str]], row_no: int) -> dict:
    if worksheet.max_row < row_no:
        return {key: "" for key, _label in columns}
    row = next(worksheet.iter_rows(min_row=row_no, max_row=row_no, values_only=True))
    return {key: normalize_cell(row[index] if index < len(row) else None) for index, (key, _label) in enumerate(columns)}


def _parse_contacts(worksheet, errors: list[dict]) -> list[dict]:
    contacts = []
    for row_no, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        values = {key: normalize_cell(row[index] if index < len(row) else None) for index, (key, _label) in enumerate(CONTACT_COLUMNS)}
        if not any(values.values()):
            continue
        if not values["name"]:
            errors.append(_import_error(CONTACT_SHEET, row_no, "姓名", "姓名不能为空"))
        if values["email"] and not EMAIL_PATTERN.match(values["email"]):
            errors.append(_import_error(CONTACT_SHEET, row_no, "电子邮件", "电子邮件格式不正确"))
        contacts.append({"id": new_id("contact"), **{key: value or None for key, value in values.items()}})
    return contacts


def _parse_date(value: str, sheet_name: str, field: str, errors: list[dict]) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = normalize_cell(value)
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    errors.append(_import_error(sheet_name, 2, field, "日期格式必须为 YYYY-MM-DD"))
    return None


def _reference_items(value: str | None, existing_by_name: dict[str, str]) -> list[dict]:
    names = [item.strip() for item in SPLIT_PATTERN.split(value or "") if item.strip()]
    result = []
    seen = set()
    for name in names:
        if name in seen:
            continue
        seen.add(name)
        result.append({"id": existing_by_name.get(name) or name, "name": name})
    return result


def _current_reference_maps(project_id: str) -> dict[str, dict[str, str]]:
    session = SessionLocal()
    basic = session.query(ProjectBasicInfo).filter_by(project_id=project_id, deleted=False).first()
    maps = {
        "laws": _map_references(getattr(basic, "laws", None)),
        "standards": _map_references(getattr(basic, "standards", None)),
    }
    custom_refs = session.query(ProjectReference).filter_by(project_id=project_id, deleted=False).all()
    for ref in custom_refs:
        if ref.type == "LAW":
            maps["laws"].setdefault(ref.name, ref.id)
        elif ref.type == "STANDARD":
            maps["standards"].setdefault(ref.name, ref.id)
    return maps


def _map_references(items) -> dict[str, str]:
    result = {}
    for item in items or []:
        if isinstance(item, dict) and item.get("name"):
            result[item["name"]] = item.get("id") or item["name"]
    return result


def _join_names(items: list) -> str:
    names = []
    for item in items:
        if isinstance(item, dict):
            name = item.get("name")
        else:
            name = str(item) if item is not None else ""
        if name:
            names.append(name)
    return "、".join(names)


def _import_error(sheet_name: str, row_no: int, field: str, reason: str) -> dict:
    return {"sheetName": sheet_name, "rowNo": row_no, "field": field, "reason": reason}
