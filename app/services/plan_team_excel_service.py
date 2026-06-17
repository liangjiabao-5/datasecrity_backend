from __future__ import annotations

from openpyxl import Workbook

from app.extensions import SessionLocal
from app.models import AssessmentTeamMember, ClientTeamMember
from app.services.audit_service import audit
from app.services.excel_common import (
    load_import_workbook,
    normalize_cell,
    raise_import_validation,
    set_column_widths,
    validate_workbook_schema,
    workbook_response,
)
from app.services.project_service import get_project


ASSESSMENT_TEAM_SHEET = "评估团队"
CLIENT_TEAM_SHEET = "被评估方团队"

ASSESSMENT_TEAM_COLUMNS = [
    ("name", "姓名"),
    ("organization", "单位"),
    ("role", "角色"),
]
CLIENT_TEAM_COLUMNS = [
    ("department", "公司/部门"),
    ("name", "姓名"),
    ("position", "职位"),
    ("contact", "联系方式"),
]

SHEET_SPECS = [
    (ASSESSMENT_TEAM_SHEET, [label for _key, label in ASSESSMENT_TEAM_COLUMNS]),
    (CLIENT_TEAM_SHEET, [label for _key, label in CLIENT_TEAM_COLUMNS]),
]


def export_template_workbook(project_id: str) -> tuple[str, bytes, str]:
    get_project(project_id)
    return workbook_response(f"评估方案团队模板-{project_id}.xlsx", _build_workbook())


def export_team_workbook(project_id: str) -> tuple[str, bytes, str]:
    get_project(project_id)
    session = SessionLocal()
    assessment_team = _active_rows(session, project_id, AssessmentTeamMember)
    client_team = _active_rows(session, project_id, ClientTeamMember)
    return workbook_response(f"评估方案团队-{project_id}.xlsx", _build_workbook(assessment_team, client_team))


def import_team_workbook(project_id: str, file) -> dict:
    get_project(project_id)
    workbook = load_import_workbook(file)
    validate_workbook_schema(workbook, SHEET_SPECS)

    errors = []
    assessment_team = _parse_assessment_team(workbook[ASSESSMENT_TEAM_SHEET], errors)
    client_team = _parse_client_team(workbook[CLIENT_TEAM_SHEET], errors)
    if errors:
        raise_import_validation(errors)

    session = SessionLocal()
    session.query(AssessmentTeamMember).filter_by(project_id=project_id, deleted=False).update({"deleted": True})
    session.query(ClientTeamMember).filter_by(project_id=project_id, deleted=False).update({"deleted": True})
    assessment_records = [AssessmentTeamMember(project_id=project_id, **row) for row in assessment_team]
    client_records = [ClientTeamMember(project_id=project_id, **row) for row in client_team]
    session.add_all([*assessment_records, *client_records])
    audit(
        "PLAN_TEAM_IMPORT",
        "Project",
        project_id,
        after={"assessmentTeamCount": len(assessment_records), "clientTeamCount": len(client_records)},
    )
    session.commit()
    return {
        "assessmentTeam": [_serialize_assessment_member(record) for record in assessment_records],
        "clientTeam": [_serialize_client_member(record) for record in client_records],
    }


def _build_workbook(
    assessment_team: list[AssessmentTeamMember] | None = None,
    client_team: list[ClientTeamMember] | None = None,
) -> Workbook:
    workbook = Workbook()
    assessment_sheet = workbook.active
    assessment_sheet.title = ASSESSMENT_TEAM_SHEET
    _append_headers(assessment_sheet, ASSESSMENT_TEAM_COLUMNS)
    client_sheet = workbook.create_sheet(CLIENT_TEAM_SHEET)
    _append_headers(client_sheet, CLIENT_TEAM_COLUMNS)

    for row in assessment_team or []:
        assessment_sheet.append([getattr(row, key) for key, _label in ASSESSMENT_TEAM_COLUMNS])
    for row in client_team or []:
        client_sheet.append([getattr(row, key) for key, _label in CLIENT_TEAM_COLUMNS])

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        set_column_widths(sheet)
    return workbook


def _append_headers(worksheet, columns: list[tuple[str, str]]) -> None:
    worksheet.append([label for _key, label in columns])


def _active_rows(session, project_id: str, model):
    return (
        session.query(model)
        .filter(model.project_id == project_id, model.deleted.is_(False))
        .order_by(model.created_at.asc())
        .all()
    )


def _parse_assessment_team(worksheet, errors: list[dict]) -> list[dict]:
    rows = []
    for row_no, values in _iter_rows(worksheet, ASSESSMENT_TEAM_COLUMNS):
        if not any(values.values()):
            continue
        if not values["name"]:
            errors.append(_import_error(ASSESSMENT_TEAM_SHEET, row_no, "姓名", "姓名不能为空"))
        rows.append(values)
    return rows


def _parse_client_team(worksheet, errors: list[dict]) -> list[dict]:
    rows = []
    for row_no, values in _iter_rows(worksheet, CLIENT_TEAM_COLUMNS):
        if not any(values.values()):
            continue
        if not values["department"] and not values["name"]:
            errors.append(_import_error(CLIENT_TEAM_SHEET, row_no, "姓名", "姓名或公司/部门至少填写一项"))
        rows.append(values)
    return rows


def _iter_rows(worksheet, columns: list[tuple[str, str]]):
    for row_no, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        yield row_no, {
            key: normalize_cell(row[index] if index < len(row) else None) or None
            for index, (key, _label) in enumerate(columns)
        }


def _import_error(sheet_name: str, row_no: int, field: str, reason: str) -> dict:
    return {"sheetName": sheet_name, "rowNo": row_no, "field": field, "reason": reason}


def _serialize_assessment_member(record: AssessmentTeamMember) -> dict:
    return {
        "id": record.id,
        "name": record.name,
        "organization": record.organization,
        "role": record.role,
    }


def _serialize_client_member(record: ClientTeamMember) -> dict:
    return {
        "id": record.id,
        "department": record.department,
        "name": record.name,
        "position": record.position,
        "contact": record.contact,
    }
