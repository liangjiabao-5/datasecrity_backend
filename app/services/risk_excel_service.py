from __future__ import annotations

import re

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.extensions import SessionLocal
from app.models import Project, ProjectRiskSummaryRecord
from app.services import harm_analysis_service
from app.services.audit_service import audit
from app.services.excel_common import load_import_workbook, normalize_cell, set_column_widths, workbook_response
from app.services.project_service import get_project
from app.services.risk_service import DEFAULT_HARM_LEVEL, DEFAULT_POSSIBILITY_LEVEL


SOURCE_COLUMNS = [
    ("index", "序号"),
    ("risk_types", "风险类型"),
    ("check_point", "检查要点"),
    ("evaluation_record", "评估结果"),
    ("risk_description", "风险描述"),
    ("risk_source_description", "风险源描述"),
    ("related_data", "涉及的数据及类型、级别"),
    ("related_activities", "涉及的数据处理活动"),
]

ITEM_COLUMNS = [
    ("index", "序号"),
    ("risk_types", "风险类型"),
    ("risk_description", "风险描述"),
    ("harm_level", "危害程度"),
    ("possibility_level", "发生可能性"),
    ("risk_source_description", "风险源描述"),
    ("risk_level", "风险等级"),
    ("related_data", "涉及的数据及类型、级别"),
    ("related_activities", "涉及的数据处理活动"),
]

SOURCE_HEADER_ALIASES = {
    "序号": "index",
    "风险类型": "risk_types",
    "riskTypes": "risk_types",
    "检查要点": "check_point",
    "checkPoint": "check_point",
    "评估结果": "evaluation_record",
    "evaluationRecord": "evaluation_record",
    "风险描述": "risk_description",
    "riskDescription": "risk_description",
    "风险源描述": "risk_source_description",
    "riskSourceDescription": "risk_source_description",
    "涉及的数据及类型、级别": "related_data",
    "relatedData": "related_data",
    "涉及的数据处理活动": "related_activities",
    "relatedActivities": "related_activities",
    "riskSourceId": "risk_source_id",
    "riskRecordId": "risk_record_id",
}

ITEM_HEADER_ALIASES = {
    "序号": "index",
    "风险类型": "risk_types",
    "riskTypes": "risk_types",
    "风险描述": "risk_description",
    "riskDescription": "risk_description",
    "危害程度": "harm_level",
    "harmLevel": "harm_level",
    "发生可能性": "possibility_level",
    "possibilityLevel": "possibility_level",
    "风险源描述": "risk_source_description",
    "riskSourceDescription": "risk_source_description",
    "风险等级": "risk_level",
    "riskLevel": "risk_level",
    "涉及的数据及类型、级别": "related_data",
    "relatedData": "related_data",
    "涉及的数据处理活动": "related_activities",
    "relatedActivities": "related_activities",
    "riskItemId": "risk_item_id",
    "riskRecordId": "risk_record_id",
}

SOURCE_EDITABLE_FIELDS = ["risk_types", "risk_description", "risk_source_description", "related_data", "related_activities"]
ITEM_EDITABLE_FIELDS = [
    "risk_types",
    "risk_description",
    "harm_level",
    "possibility_level",
    "risk_source_description",
    "related_data",
    "related_activities",
]
SOURCE_READONLY_FIELDS = [("check_point", "检查要点"), ("evaluation_record", "评估结果")]

MULTI_VALUE_PATTERN = re.compile(r"[、,，;；\r\n]+")
EMPTY_LEVEL_TEXTS = {"", "-", "待评估", "TO_BE_EVALUATED", "PENDING"}
HARM_LEVEL_ALIASES = {
    "MAJOR": "VERY_HIGH",
    "VERY_HIGH": "VERY_HIGH",
    "很高": "VERY_HIGH",
    "HIGH": "HIGH",
    "高": "HIGH",
    "RELATIVELY_HIGH": "RELATIVELY_HIGH",
    "较高": "RELATIVELY_HIGH",
    "MEDIUM": "MEDIUM",
    "中": "MEDIUM",
    "LOW": "LOW",
    "低": "LOW",
    "SLIGHT": "LOW",
}
POSSIBILITY_LEVEL_ALIASES = {
    "HIGH": "HIGH",
    "高": "HIGH",
    "MEDIUM": "MEDIUM",
    "中": "MEDIUM",
    "LOW": "LOW",
    "低": "LOW",
}
HARM_LEVEL_NAMES = {
    "VERY_HIGH": "很高",
    "HIGH": "高",
    "RELATIVELY_HIGH": "较高",
    "MEDIUM": "中",
    "LOW": "低",
    DEFAULT_HARM_LEVEL: "待评估",
}
POSSIBILITY_LEVEL_NAMES = {
    "HIGH": "高",
    "MEDIUM": "中",
    "LOW": "低",
    DEFAULT_POSSIBILITY_LEVEL: "待评估",
}
RISK_LEVEL_NAMES = {
    "MAJOR": "重大安全风险",
    "HIGH": "高安全风险",
    "MEDIUM": "中安全风险",
    "LOW": "低安全风险",
    "SLIGHT": "轻微安全风险",
}


def export_risk_sources_workbook(project_id: str) -> tuple[str, bytes, str]:
    get_project(project_id)
    records = _current_records(project_id)
    workbook = _build_workbook(records, SOURCE_COLUMNS, "riskSourceId")
    return workbook_response(f"数据安全风险源清单-{project_id}.xlsx", workbook)


def export_risk_items_workbook(project_id: str) -> tuple[str, bytes, str]:
    get_project(project_id)
    records = _current_records(project_id)
    workbook = _build_workbook(records, ITEM_COLUMNS, "riskItemId")
    return workbook_response(f"数据安全风险清单-{project_id}.xlsx", workbook)


def import_risk_sources_workbook(project_id: str, file) -> dict:
    get_project(project_id)
    workbook = load_import_workbook(file)
    worksheet = workbook.active
    header_map = _header_map(worksheet, SOURCE_HEADER_ALIASES)

    session = SessionLocal()
    records_by_id = _current_records_by_id(session, project_id)
    imported = 0
    errors = []
    seen_record_ids = set()

    for row_no, row in _iter_import_rows(worksheet):
        record, locator_field = _locate_record(row, header_map, records_by_id, "risk_source_id")
        if not record:
            errors.append(_import_error(row_no, locator_field, "无法定位对应风险记录，请使用最新导出模板。"))
            continue
        if record.id in seen_record_ids:
            errors.append(_import_error(row_no, locator_field, "同一导入文件中已出现相同风险记录。"))
            continue
        seen_record_ids.add(record.id)

        row_errors = _validate_source_readonly_fields(record, row, header_map, row_no)
        if row_errors:
            errors.extend(row_errors)
            continue

        _apply_source_values(record, row, header_map)
        record.manual_adjusted = True
        imported += 1

    audit("RISK_SOURCE_IMPORT", "Project", project_id, after={"imported": imported, "failed": len(errors)})
    session.commit()
    return {"importedCount": imported, "failedCount": len(errors), "errors": errors}


def import_risk_items_workbook(project_id: str, file) -> dict:
    project = get_project(project_id)
    workbook = load_import_workbook(file)
    worksheet = workbook.active
    header_map = _header_map(worksheet, ITEM_HEADER_ALIASES)

    session = SessionLocal()
    records_by_id = _current_records_by_id(session, project_id)
    imported = 0
    errors = []
    seen_record_ids = set()

    for row_no, row in _iter_import_rows(worksheet):
        record, locator_field = _locate_record(row, header_map, records_by_id, "risk_item_id")
        if not record:
            errors.append(_import_error(row_no, locator_field, "无法定位对应风险记录，请使用最新导出模板。"))
            continue
        if record.id in seen_record_ids:
            errors.append(_import_error(row_no, locator_field, "同一导入文件中已出现相同风险记录。"))
            continue
        seen_record_ids.add(record.id)

        row_errors = _apply_item_values(session, project, record, row, header_map, row_no)
        if row_errors:
            errors.extend(row_errors)
            continue
        record.manual_adjusted = True
        imported += 1

    audit("RISK_ITEM_IMPORT", "Project", project_id, after={"imported": imported, "failed": len(errors)})
    session.commit()
    return {"importedCount": imported, "failedCount": len(errors), "errors": errors}


def _build_workbook(records: list[ProjectRiskSummaryRecord], columns: list[tuple[str, str]], id_header: str) -> Workbook:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "汇总分析"
    worksheet.append([label for _key, label in columns] + [id_header])
    for index, record in enumerate(records, start=1):
        row = []
        for key, _label in columns:
            row.append(_export_value(record, key, index))
        row.append(record.id)
        worksheet.append(row)
    worksheet.freeze_panes = "A2"
    set_column_widths(worksheet)
    id_column_letter = get_column_letter(len(columns) + 1)
    worksheet.column_dimensions[id_column_letter].hidden = True
    return workbook


def _export_value(record: ProjectRiskSummaryRecord, key: str, index: int):
    if key == "index":
        return index
    value = getattr(record, key)
    if key in {"risk_types", "related_activities"}:
        return "、".join(str(item) for item in (value or []))
    if key == "harm_level":
        return HARM_LEVEL_NAMES.get(value, value)
    if key == "possibility_level":
        return POSSIBILITY_LEVEL_NAMES.get(value, value)
    if key == "risk_level":
        return RISK_LEVEL_NAMES.get(value, value)
    return value


def _current_records(project_id: str) -> list[ProjectRiskSummaryRecord]:
    session = SessionLocal()
    return (
        session.query(ProjectRiskSummaryRecord)
        .filter(
            ProjectRiskSummaryRecord.project_id == project_id,
            ProjectRiskSummaryRecord.deleted.is_(False),
            ProjectRiskSummaryRecord.current.is_(True),
        )
        .order_by(ProjectRiskSummaryRecord.created_at.asc())
        .all()
    )


def _current_records_by_id(session, project_id: str) -> dict[str, ProjectRiskSummaryRecord]:
    records = (
        session.query(ProjectRiskSummaryRecord)
        .filter(
            ProjectRiskSummaryRecord.project_id == project_id,
            ProjectRiskSummaryRecord.deleted.is_(False),
            ProjectRiskSummaryRecord.current.is_(True),
        )
        .order_by(ProjectRiskSummaryRecord.created_at.asc())
        .all()
    )
    return {record.id: record for record in records}


def _header_map(worksheet, aliases: dict[str, str]) -> dict[str, int]:
    result = {}
    for index, cell in enumerate(worksheet[1]):
        key = aliases.get(normalize_cell(cell.value))
        if key and key not in result:
            result[key] = index
    return result


def _iter_import_rows(worksheet):
    for row_no, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(normalize_cell(value) for value in row):
            continue
        yield row_no, row


def _locate_record(
    row: tuple,
    header_map: dict[str, int],
    records_by_id: dict[str, ProjectRiskSummaryRecord],
    preferred_id_key: str,
) -> tuple[ProjectRiskSummaryRecord | None, str]:
    record_id = _cell_text(row, header_map, preferred_id_key)
    locator_field = _field_label(preferred_id_key)
    if record_id:
        return records_by_id.get(record_id), locator_field

    record_id = _cell_text(row, header_map, "risk_record_id")
    if record_id:
        return records_by_id.get(record_id), "riskRecordId"

    matches = list(records_by_id.values())
    check_point = _cell_text(row, header_map, "check_point")
    evaluation_record = _cell_text(row, header_map, "evaluation_record")
    has_context_locator = False
    if check_point:
        matches = [record for record in matches if _same_text(record.check_point, check_point)]
        locator_field = "检查要点"
        has_context_locator = True
    if evaluation_record:
        matches = [record for record in matches if _same_text(record.evaluation_record, evaluation_record)]
        locator_field = "评估结果"
        has_context_locator = True
    if not has_context_locator:
        return None, locator_field
    return (matches[0], locator_field) if len(matches) == 1 else (None, locator_field)


def _validate_source_readonly_fields(
    record: ProjectRiskSummaryRecord,
    row: tuple,
    header_map: dict[str, int],
    row_no: int,
) -> list[dict]:
    errors = []
    for key, label in SOURCE_READONLY_FIELDS:
        if key not in header_map:
            continue
        value = _cell_text(row, header_map, key)
        if not _same_text(getattr(record, key), value):
            errors.append(_import_error(row_no, label, f"{label}与系统风险记录不一致，请使用最新导出模板。"))
    return errors


def _apply_source_values(record: ProjectRiskSummaryRecord, row: tuple, header_map: dict[str, int]) -> None:
    for field in SOURCE_EDITABLE_FIELDS:
        if field not in header_map:
            continue
        if field in {"risk_types", "related_activities"}:
            setattr(record, field, _split_multi_value(_cell_text(row, header_map, field)))
        else:
            setattr(record, field, _cell_text(row, header_map, field) or None)


def _apply_item_values(
    session,
    project: Project,
    record: ProjectRiskSummaryRecord,
    row: tuple,
    header_map: dict[str, int],
    row_no: int,
) -> list[dict]:
    errors = []
    normalized_harm = None
    normalized_possibility = None
    if "harm_level" in header_map:
        normalized_harm = _normalize_harm_level(_cell_text(row, header_map, "harm_level"))
        if normalized_harm is None:
            errors.append(_import_error(row_no, "危害程度", "危害程度只能填写很高、高、较高、中、低、待评估或对应枚举值。"))
    if "possibility_level" in header_map:
        normalized_possibility = _normalize_possibility_level(_cell_text(row, header_map, "possibility_level"))
        if normalized_possibility is None:
            errors.append(_import_error(row_no, "发生可能性", "发生可能性只能填写高、中、低、待评估或对应枚举值。"))
    if errors:
        return errors

    for field in ITEM_EDITABLE_FIELDS:
        if field not in header_map:
            continue
        if field in {"risk_types", "related_activities"}:
            setattr(record, field, _split_multi_value(_cell_text(row, header_map, field)))
        elif field == "harm_level":
            record.harm_level = normalized_harm
        elif field == "possibility_level":
            record.possibility_level = normalized_possibility
        else:
            setattr(record, field, _cell_text(row, header_map, field) or None)

    if "harm_level" in header_map or "possibility_level" in header_map:
        if record.harm_level in (None, "", DEFAULT_HARM_LEVEL) or record.possibility_level in (None, "", DEFAULT_POSSIBILITY_LEVEL):
            record.risk_level = None
        else:
            record.risk_level = harm_analysis_service.risk_level_from_project_matrix(
                session,
                project,
                record.harm_level,
                record.possibility_level,
            )
    return []


def _cell_text(row: tuple, header_map: dict[str, int], key: str) -> str:
    index = header_map.get(key)
    if index is None or index >= len(row):
        return ""
    return normalize_cell(row[index])


def _split_multi_value(value: str) -> list[str]:
    result = []
    seen = set()
    for item in MULTI_VALUE_PATTERN.split(value or ""):
        normalized = item.strip()
        if normalized and normalized not in seen:
            seen.add(normalized)
            result.append(normalized)
    return result


def _normalize_harm_level(value: str) -> str | None:
    normalized = _normalize_level_text(value)
    if normalized in EMPTY_LEVEL_TEXTS:
        return DEFAULT_HARM_LEVEL
    return HARM_LEVEL_ALIASES.get(normalized)


def _normalize_possibility_level(value: str) -> str | None:
    normalized = _normalize_level_text(value)
    if normalized in EMPTY_LEVEL_TEXTS:
        return DEFAULT_POSSIBILITY_LEVEL
    return POSSIBILITY_LEVEL_ALIASES.get(normalized)


def _normalize_level_text(value: str) -> str:
    text = normalize_cell(value)
    return text.upper() if re.fullmatch(r"[A-Za-z_]+", text) else text


def _same_text(left, right) -> bool:
    return " ".join(normalize_cell(left).split()) == " ".join(normalize_cell(right).split())


def _field_label(key: str) -> str:
    return {"risk_source_id": "riskSourceId", "risk_item_id": "riskItemId", "risk_record_id": "riskRecordId"}.get(key, key)


def _import_error(row_no: int, field: str, reason: str) -> dict:
    return {"rowNo": row_no, "field": _field_label(field), "reason": reason}
