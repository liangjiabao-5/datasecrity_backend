from io import BytesIO

from openpyxl import Workbook, load_workbook
from sqlalchemy import and_, or_

from app.common.exceptions import BusinessError, NotFoundError
from app.common.pagination import page_args
from app.extensions import SessionLocal
from app.models import (
    EvaluationRecord,
    ProjectAssessmentItem,
    ScoreCalculationRecord,
    ScoreModel,
    ScoreModelRange,
)
from app.services.audit_service import audit
from app.services.project_service import get_project


RESULT_NAMES = {
    "COMPLIANT": "符合",
    "PARTIAL": "基本符合",
    "NON_COMPLIANT": "不符合",
    "NOT_APPLICABLE": "不适用",
}

VALID_EVALUATION_RESULTS = set(RESULT_NAMES)
RESULT_FILTER_ERROR_MESSAGE = "符合情况筛选值只能为 COMPLIANT、PARTIAL、NON_COMPLIANT、NOT_APPLICABLE"


POSSIBILITY_NAMES = {"HIGH": "高", "MEDIUM": "中", "LOW": "低"}


SCORE_RESULT_DEFAULTS = {
    "COMPLIANT": 1.0,
    "PARTIAL": 0.5,
    "NON_COMPLIANT": 0.0,
}


SCORE_COUNT_KEYS = {
    "COMPLIANT": "compliantCount",
    "PARTIAL": "partialCount",
    "NON_COMPLIANT": "nonCompliantCount",
}


EVALUATION_EXPORT_COLUMNS = [
    ("itemId", "检查项ID"),
    ("itemCode", "检查项编号"),
    ("sheetName", "工作表"),
    ("category", "一级分类"),
    ("subcategory", "二级分类"),
    ("checkPoint", "检查要点"),
    ("evaluationRecord", "评估结果"),
    ("evaluationResult", "符合情况"),
]

RESULT_ALIASES = {
    "COMPLIANT": "COMPLIANT",
    "符合": "COMPLIANT",
    "PARTIAL": "PARTIAL",
    "基本符合": "PARTIAL",
    "NON_COMPLIANT": "NON_COMPLIANT",
    "不符合": "NON_COMPLIANT",
    "NOT_APPLICABLE": "NOT_APPLICABLE",
    "不适用": "NOT_APPLICABLE",
}

IMPORT_FIELD_LABELS = {
    "item_id": "检查项ID",
    "item_code": "检查项编号",
    "sheet_name": "工作表",
    "category": "一级分类",
    "subcategory": "二级分类",
    "check_point": "检查要点",
    "evaluation_result": "符合情况",
    "evaluation_record": "评估结果",
}

IMPORT_READONLY_FIELD_ATTRS = [
    ("item_id", "id"),
    ("item_code", "item_code"),
    ("sheet_name", "sheet_name"),
    ("category", "category"),
    ("subcategory", "subcategory"),
    ("check_point", "check_point"),
]

RESULT_IMPORT_ERROR_REASON = (
    "符合情况列只能填写符合、基本符合、不符合、不适用或对应枚举值 "
    "COMPLIANT、PARTIAL、NON_COMPLIANT、NOT_APPLICABLE。"
)


def catalog(project_id: str) -> list[dict]:
    get_project(project_id)
    session = SessionLocal()
    rows = (
        session.query(ProjectAssessmentItem)
        .filter(ProjectAssessmentItem.project_id == project_id, ProjectAssessmentItem.deleted.is_(False))
        .order_by(ProjectAssessmentItem.sort_order.asc())
        .all()
    )
    tree: dict[str, dict] = {}
    for row in rows:
        sheet = tree.setdefault(row.sheet_name or "-", {"id": row.sheet_name or "-", "name": row.sheet_name or "-", "children": {}})
        category = sheet["children"].setdefault(
            row.category or "-",
            {"id": f"{row.sheet_name}|{row.category}", "name": row.category or "-", "children": {}},
        )
        category["children"].setdefault(
            row.subcategory or "-",
            {"id": row.category_id, "name": row.subcategory or "-", "children": []},
        )

    return [_node_to_list(node) for node in tree.values()]


def list_items(project_id: str, args) -> dict:
    get_project(project_id)
    session = SessionLocal()
    result_filters = _normalize_result_filter_args(args)
    query = session.query(ProjectAssessmentItem).filter(
        ProjectAssessmentItem.project_id == project_id,
        ProjectAssessmentItem.deleted.is_(False),
    )
    if args.get("categoryId"):
        query = query.filter(ProjectAssessmentItem.category_id == args.get("categoryId"))
    sheet_name = _arg_text(args, "sheet_name") or _arg_text(args, "sheetName")
    if sheet_name:
        query = query.filter(ProjectAssessmentItem.sheet_name == sheet_name)
    category = _arg_text(args, "category")
    if category:
        query = query.filter(ProjectAssessmentItem.category == category)
    keyword = _arg_text(args, "keyword")
    if keyword:
        like_keyword = f"%{keyword}%"
        query = query.filter(
            or_(
                ProjectAssessmentItem.item_code.like(like_keyword),
                ProjectAssessmentItem.check_point.like(like_keyword),
                ProjectAssessmentItem.check_content.like(like_keyword),
            )
        )
    if result_filters:
        query = query.join(
            EvaluationRecord,
            and_(
                EvaluationRecord.project_id == project_id,
                EvaluationRecord.item_id == ProjectAssessmentItem.id,
                EvaluationRecord.deleted.is_(False),
            ),
        ).filter(EvaluationRecord.evaluation_result.in_(result_filters))

    query = query.order_by(ProjectAssessmentItem.sort_order.asc())
    page_no, page_size = page_args()
    total = query.count()
    rows = query.offset((page_no - 1) * page_size).limit(page_size).all()
    records = _records_by_item(session, project_id, [row.id for row in rows])
    return {
        "list": [_serialize_item(row, records.get(row.id)) for row in rows],
        "pageNo": page_no,
        "pageSize": page_size,
        "total": total,
    }


def _arg_text(args, key: str) -> str | None:
    value = args.get(key)
    if value in (None, ""):
        return None
    text = str(value).strip()
    return text or None


def _normalize_result_filter_args(args) -> list[str]:
    raw_results = []
    raw_results.extend(_arg_values(args, "results[]"))
    raw_results.extend(_arg_values(args, "results"))

    results = _split_result_filter_values(raw_results)
    if not results:
        results = _split_result_filter_values(_arg_values(args, "result"))

    results = list(dict.fromkeys(results))
    invalid_results = [value for value in results if value not in VALID_EVALUATION_RESULTS]
    if invalid_results:
        raise BusinessError("INVALID_EVALUATION_RESULT_FILTER", RESULT_FILTER_ERROR_MESSAGE)
    return results


def _arg_values(args, key: str) -> list:
    if hasattr(args, "getlist"):
        return args.getlist(key)
    value = args.get(key) if hasattr(args, "get") else None
    if isinstance(value, (list, tuple)):
        return list(value)
    return [] if value is None else [value]


def _split_result_filter_values(values: list) -> list[str]:
    results = []
    for value in values:
        if value in (None, ""):
            continue
        for item in str(value).split(","):
            result = item.strip()
            if result:
                results.append(result)
    return results


def save_record(project_id: str, item_id: str, payload: dict) -> dict:
    get_project(project_id)
    session = SessionLocal()
    item = _get_item(session, project_id, item_id)
    record = (
        session.query(EvaluationRecord)
        .filter(EvaluationRecord.project_id == project_id, EvaluationRecord.item_id == item.id, EvaluationRecord.deleted.is_(False))
        .first()
    )
    if not record:
        record = EvaluationRecord(project_id=project_id, item_id=item.id)
        session.add(record)
    for field in ["evaluation_result", "evaluation_record"]:
        if field in payload:
            setattr(record, field, payload.get(field))
    record.manual_updated = True
    audit("EVALUATION_RECORD_SAVE", "EvaluationRecord", record.id, after=record.to_dict())
    session.commit()
    return _serialize_item(item, record)


def batch_result(project_id: str, payload: dict) -> dict:
    get_project(project_id)
    session = SessionLocal()
    item_ids = payload.get("item_ids") or []
    result = payload.get("evaluation_result")
    updated = 0
    for item_id in item_ids:
        item = _get_item(session, project_id, item_id)
        record = (
            session.query(EvaluationRecord)
            .filter(EvaluationRecord.project_id == project_id, EvaluationRecord.item_id == item.id, EvaluationRecord.deleted.is_(False))
            .first()
        )
        if not record:
            record = EvaluationRecord(project_id=project_id, item_id=item.id)
            session.add(record)
        record.evaluation_result = result
        record.manual_updated = True
        updated += 1
    audit("EVALUATION_BATCH_RESULT", "Project", project_id, after={"updated": updated, "result": result})
    session.commit()
    return {"updated": updated}


def calculate_score(project_id: str) -> dict:
    project = get_project(project_id)
    session = SessionLocal()
    calculation = calculate_project_score_snapshot(session, project)
    record = ScoreCalculationRecord(
        project_id=project_id,
        score=calculation["score"],
        possibility_level=calculation["possibilityLevel"],
        score_model_id=calculation["scoreModelId"],
        score_model_version=calculation["scoreModelVersion"],
        calculation_detail=calculation["detail"],
    )
    session.add(record)
    audit("SCORE_CALCULATE", "Project", project_id, after={"score": calculation["score"], "possibilityLevel": calculation["possibilityLevel"]})
    session.commit()
    return {
        "score": calculation["score"],
        "possibilityLevel": calculation["possibilityLevel"],
        "possibilityLevelName": POSSIBILITY_NAMES.get(calculation["possibilityLevel"], calculation["possibilityLevel"]),
        "scoreModelVersion": record.score_model_version,
        "detail": calculation["detail"],
    }


def calculate_project_score_snapshot(session, project) -> dict:
    rows = (
        session.query(EvaluationRecord)
        .filter(EvaluationRecord.project_id == project.id, EvaluationRecord.deleted.is_(False))
        .all()
    )
    counts = {
        "compliantCount": 0,
        "partialCount": 0,
        "nonCompliantCount": 0,
        "notApplicableCount": 0,
    }
    for row in rows:
        if row.evaluation_result == "COMPLIANT":
            counts["compliantCount"] += 1
        elif row.evaluation_result == "PARTIAL":
            counts["partialCount"] += 1
        elif row.evaluation_result == "NON_COMPLIANT":
            counts["nonCompliantCount"] += 1
        elif row.evaluation_result == "NOT_APPLICABLE":
            counts["notApplicableCount"] += 1

    score_model = session.get(ScoreModel, project.score_model_id) if project.score_model_id else None
    denominator = counts["compliantCount"] + counts["partialCount"] + counts["nonCompliantCount"]
    numerator = sum(
        counts[count_key] * _result_score(score_model, result, SCORE_RESULT_DEFAULTS[result])
        for result, count_key in SCORE_COUNT_KEYS.items()
    )
    score = 0.0 if denominator == 0 else round(numerator / denominator * 100, 2)
    level = possibility_level_for_score(session, project.score_model_id, score)
    return {
        "score": score,
        "possibilityLevel": level,
        "scoreModelId": project.score_model_id,
        "scoreModelVersion": score_model.version if score_model else 1,
        "detail": counts,
    }


def _result_score(score_model: ScoreModel | None, result: str, default: float) -> float:
    if not score_model or not isinstance(score_model.result_scores, dict):
        return default
    value = score_model.result_scores.get(result)
    if value is None:
        value = score_model.result_scores.get(result.lower())
    if value is None:
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def export_template_workbook(project_id: str) -> tuple[str, bytes, str]:
    get_project(project_id)
    rows = _all_items_with_records(project_id)
    workbook = _build_workbook(rows, include_records=False)
    return _workbook_response(f"{project_id}_evaluation_template.xlsx", workbook)


def export_records_workbook(project_id: str) -> tuple[str, bytes, str]:
    get_project(project_id)
    rows = _all_items_with_records(project_id)
    workbook = _build_workbook(rows, include_records=True)
    return _workbook_response(f"{project_id}_evaluation_records.xlsx", workbook)


def import_records_workbook(project_id: str, file) -> dict:
    get_project(project_id)
    if not file or not file.filename:
        raise BusinessError("FILE_REQUIRED", "Upload file is required.")
    try:
        file.stream.seek(0)
        workbook = load_workbook(file.stream, data_only=True)
    except Exception as exc:
        raise BusinessError("INVALID_IMPORT_FILE", "Only valid xlsx files are supported.") from exc

    worksheet = workbook.active
    headers = [_normalize_header(cell.value) for cell in worksheet[1]]
    column_map = {_header_to_key(header): index for index, header in enumerate(headers) if _header_to_key(header)}
    if "item_id" not in column_map and "item_code" not in column_map:
        raise BusinessError("INVALID_IMPORT_FILE", "Import file must include itemId or itemCode.")

    session = SessionLocal()
    imported = 0
    errors = []
    for row_no, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(value not in (None, "") for value in row):
            continue
        item = _item_from_import_row(session, project_id, row, column_map)
        if not item:
            errors.append(_import_error(row_no, "item_id", "未找到对应检查项，请使用最新导出模板。"))
            continue
        result = _cell_value(row, column_map, "evaluation_result")
        editable_values = [result, _cell_value(row, column_map, "evaluation_record")]
        if not any(value not in (None, "") for value in editable_values):
            continue
        row_errors = _validate_import_readonly_fields(item, row, column_map, row_no)
        normalized_result = _normalize_result(result)
        if result and not normalized_result:
            row_errors.append(_import_error(row_no, "evaluation_result", RESULT_IMPORT_ERROR_REASON))
        if row_errors:
            errors.extend(row_errors)
            continue
        record = _get_or_create_record(session, project_id, item.id)
        record.evaluation_result = normalized_result
        record.evaluation_record = _cell_value(row, column_map, "evaluation_record")
        record.manual_updated = True
        audit("EVALUATION_RECORD_SAVE", "EvaluationRecord", record.id, after=record.to_dict())
        imported += 1

    audit("EVALUATION_IMPORT", "Project", project_id, after={"imported": imported, "failed": len(errors)})
    session.commit()
    return {"projectId": project_id, "importedCount": imported, "failedCount": len(errors), "errors": errors}


def _all_items_with_records(project_id: str) -> list[tuple[ProjectAssessmentItem, EvaluationRecord | None]]:
    session = SessionLocal()
    items = (
        session.query(ProjectAssessmentItem)
        .filter(ProjectAssessmentItem.project_id == project_id, ProjectAssessmentItem.deleted.is_(False))
        .order_by(ProjectAssessmentItem.sort_order.asc())
        .all()
    )
    records = _records_by_item(session, project_id, [item.id for item in items])
    return [(item, records.get(item.id)) for item in items]


def _build_workbook(rows: list[tuple[ProjectAssessmentItem, EvaluationRecord | None]], include_records: bool) -> Workbook:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "现场测评"
    worksheet.append([label for _, label in EVALUATION_EXPORT_COLUMNS])
    for item, record in rows:
        serialized = _serialize_item(item, record if include_records else None)
        values = []
        for key, _label in EVALUATION_EXPORT_COLUMNS:
            value = serialized.get(key)
            values.append(value)
        worksheet.append(values)
    worksheet.freeze_panes = "A2"
    for column_cells in worksheet.columns:
        column_letter = column_cells[0].column_letter
        worksheet.column_dimensions[column_letter].width = min(max(len(str(column_cells[0].value or "")) + 4, 16), 45)
    return workbook


def _workbook_response(file_name: str, workbook: Workbook) -> tuple[str, bytes, str]:
    stream = BytesIO()
    workbook.save(stream)
    return file_name, stream.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _normalize_header(value) -> str:
    return str(value or "").strip()


def _header_to_key(header: str) -> str | None:
    aliases = {
        "itemId": "item_id",
        "检查项ID": "item_id",
        "itemCode": "item_code",
        "检查项编号": "item_code",
        "sheetName": "sheet_name",
        "工作表": "sheet_name",
        "category": "category",
        "一级分类": "category",
        "subcategory": "subcategory",
        "二级分类": "subcategory",
        "checkPoint": "check_point",
        "检查要点": "check_point",
        "evaluationResult": "evaluation_result",
        "测评结果": "evaluation_result",
        "符合情况": "evaluation_result",
        "evaluationRecord": "evaluation_record",
        "测评记录": "evaluation_record",
        "评估结果": "evaluation_record",
    }
    return aliases.get(header)


def _item_from_import_row(session, project_id: str, row: tuple, column_map: dict[str, int]) -> ProjectAssessmentItem | None:
    item_id = _import_text(_cell_value(row, column_map, "item_id"))
    if item_id:
        item = (
            session.query(ProjectAssessmentItem)
            .filter(
                ProjectAssessmentItem.project_id == project_id,
                ProjectAssessmentItem.id == item_id,
                ProjectAssessmentItem.deleted.is_(False),
            )
            .first()
        )
        if item:
            return item
    item_code = _import_text(_cell_value(row, column_map, "item_code"))
    if item_code:
        candidates = (
            session.query(ProjectAssessmentItem)
            .filter(
                ProjectAssessmentItem.project_id == project_id,
                ProjectAssessmentItem.item_code == item_code,
                ProjectAssessmentItem.deleted.is_(False),
            )
            .all()
        )
        return _match_item_by_import_context(candidates, row, column_map)
    return None


def _match_item_by_import_context(
    candidates: list[ProjectAssessmentItem],
    row: tuple,
    column_map: dict[str, int],
) -> ProjectAssessmentItem | None:
    if len(candidates) <= 1:
        return candidates[0] if candidates else None

    matches = candidates
    for key, attr in [
        ("sheet_name", "sheet_name"),
        ("category", "category"),
        ("subcategory", "subcategory"),
        ("check_point", "check_point"),
    ]:
        value = _import_text(_cell_value(row, column_map, key))
        if not value:
            continue
        narrowed = [item for item in matches if _import_text(getattr(item, attr)) == value]
        if narrowed:
            matches = narrowed
        if len(matches) == 1:
            return matches[0]
    return matches[0] if len(matches) == 1 else None


def _validate_import_readonly_fields(
    item: ProjectAssessmentItem,
    row: tuple,
    column_map: dict[str, int],
    row_no: int,
) -> list[dict]:
    errors = []
    for key, attr in IMPORT_READONLY_FIELD_ATTRS:
        if key not in column_map:
            continue
        value = _cell_value(row, column_map, key)
        if key == "item_id" and value in (None, ""):
            continue
        expected = getattr(item, attr)
        if _import_text(value) != _import_text(expected):
            field_label = IMPORT_FIELD_LABELS[key]
            errors.append(_import_error(row_no, key, f"{field_label}与系统检查项不一致，请使用最新导出模板。"))
    return errors


def _import_error(row_no: int, field_key: str, reason: str) -> dict:
    return {"rowNo": row_no, "field": IMPORT_FIELD_LABELS.get(field_key, field_key), "reason": reason}


def _cell_value(row: tuple, column_map: dict[str, int], key: str):
    index = column_map.get(key)
    if index is None or index >= len(row):
        return None
    value = row[index]
    if isinstance(value, str):
        value = value.strip()
    return value


def _import_text(value) -> str:
    if value in (None, ""):
        return ""
    return " ".join(str(value).strip().split())


def _normalize_result(value) -> str | None:
    if value in (None, ""):
        return None
    return RESULT_ALIASES.get(str(value).strip())


def _records_by_item(session, project_id: str, item_ids: list[str]) -> dict[str, EvaluationRecord]:
    if not item_ids:
        return {}
    rows = (
        session.query(EvaluationRecord)
        .filter(
            EvaluationRecord.project_id == project_id,
            EvaluationRecord.item_id.in_(item_ids),
            EvaluationRecord.deleted.is_(False),
        )
        .all()
    )
    return {row.item_id: row for row in rows}


def _get_or_create_record(session, project_id: str, item_id: str) -> EvaluationRecord:
    record = (
        session.query(EvaluationRecord)
        .filter(
            EvaluationRecord.project_id == project_id,
            EvaluationRecord.item_id == item_id,
            EvaluationRecord.deleted.is_(False),
        )
        .first()
    )
    if not record:
        record = EvaluationRecord(project_id=project_id, item_id=item_id)
        session.add(record)
    return record


def _serialize_item(item: ProjectAssessmentItem, record: EvaluationRecord | None) -> dict:
    data = item.to_dict()
    data["itemId"] = item.id
    data["evaluationResult"] = record.evaluation_result if record else None
    data["evaluationResultName"] = RESULT_NAMES.get(record.evaluation_result) if record else None
    data["evaluationRecord"] = record.evaluation_record if record else None
    return data


def _get_item(session, project_id: str, item_id: str) -> ProjectAssessmentItem:
    item = (
        session.query(ProjectAssessmentItem)
        .filter(ProjectAssessmentItem.project_id == project_id, ProjectAssessmentItem.id == item_id, ProjectAssessmentItem.deleted.is_(False))
        .first()
    )
    if not item:
        raise NotFoundError("Assessment item not found.")
    return item


def possibility_level_for_score(session, score_model_id: str | None, score: float) -> str:
    ranges = []
    if score_model_id:
        ranges = (
            session.query(ScoreModelRange)
            .filter(ScoreModelRange.score_model_id == score_model_id, ScoreModelRange.deleted.is_(False))
            .order_by(ScoreModelRange.min_score.asc())
            .all()
        )
    for item in ranges:
        if score >= item.min_score and (score < item.max_score or (score == 100 and item.max_score == 100)):
            return item.level
    if score < 60:
        return "HIGH"
    if score < 80:
        return "MEDIUM"
    return "LOW"


def _node_to_list(node: dict) -> dict:
    children = node.get("children", {})
    return {
        "id": node["id"],
        "name": node["name"],
        "children": [_node_to_list(child) for child in children.values()] if isinstance(children, dict) else children,
    }
