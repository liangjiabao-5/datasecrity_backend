from __future__ import annotations

from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from app.common.exceptions import BusinessError


EXCEL_MIME_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
ALLOWED_EXCEL_EXTENSIONS = {".xlsx", ".xls"}


def workbook_response(file_name: str, workbook) -> tuple[str, bytes, str]:
    stream = BytesIO()
    workbook.save(stream)
    return file_name, stream.getvalue(), EXCEL_MIME_TYPE


def load_import_workbook(file):
    if not file or not file.filename:
        raise BusinessError("FILE_REQUIRED", "Upload file is required.")
    suffix = Path(file.filename).suffix.lower()
    if suffix not in ALLOWED_EXCEL_EXTENSIONS:
        raise_import_validation(
            [
                {
                    "sheetName": None,
                    "rowNo": 0,
                    "field": "file",
                    "reason": "导入文件必须是 .xlsx 或 .xls 格式。",
                }
            ]
        )
    try:
        return load_workbook_from_upload(file)
    except Exception as exc:
        raise_import_validation(
            [
                {
                    "sheetName": None,
                    "rowNo": 0,
                    "field": "file",
                    "reason": "导入文件不是有效的 Excel 文件。",
                }
            ],
            exc,
        )


def load_workbook_from_upload(file):
    file.stream.seek(0)
    content = file.read()
    return load_workbook(BytesIO(content), data_only=True)


def validate_workbook_schema(workbook, sheet_specs: list[tuple[str, list[str]]]) -> None:
    errors = []
    for sheet_name, expected_headers in sheet_specs:
        if sheet_name not in workbook.sheetnames:
            errors.append(
                {
                    "sheetName": sheet_name,
                    "rowNo": 0,
                    "field": sheet_name,
                    "reason": "缺少必需工作表，请使用最新模板。",
                }
            )
            continue
        actual_headers = [normalize_cell(cell.value) for cell in workbook[sheet_name][1][: len(expected_headers)]]
        for index, expected in enumerate(expected_headers):
            actual = actual_headers[index] if index < len(actual_headers) else ""
            if actual != expected:
                errors.append(
                    {
                        "sheetName": sheet_name,
                        "rowNo": 1,
                        "field": expected,
                        "reason": "表头名称与模板不一致，请使用最新模板。",
                    }
                )
    if errors:
        raise_import_validation(errors)


def raise_import_validation(errors: list[dict], cause: Exception | None = None):
    error = BusinessError(
        "IMPORT_VALIDATION_FAILED",
        "导入文件存在格式错误",
        data={"errors": errors},
    )
    if cause:
        raise error from cause
    raise error


def normalize_cell(value) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def set_column_widths(worksheet, min_width: int = 14, max_width: int = 45) -> None:
    for column_cells in worksheet.columns:
        header = column_cells[0]
        width = min(max(len(str(header.value or "")) + 4, min_width), max_width)
        worksheet.column_dimensions[header.column_letter].width = width
