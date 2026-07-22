import hashlib
import re
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import text

from app.extensions import SessionLocal
from app.models import (
    AssessmentTemplate,
    AssessmentTemplateItem,
    HarmModel,
    HarmModelRule,
    ProjectAssessmentItem,
    ProjectRiskSummaryRecord,
    RemediationSuggestionTemplate,
    RiskMatrix,
    RiskSourceTemplate,
    ScoreModel,
    ScoreModelRange,
)


DEFAULT_SCORE_RANGES = [
    ("HIGH", 0, 60, True, False),
    ("MEDIUM", 60, 80, True, False),
    ("LOW", 80, 100, True, True),
]


DEFAULT_RISK_MATRIX = {
    "HIGH": {
        "VERY_HIGH": "MAJOR",
        "HIGH": "MAJOR",
        "RELATIVELY_HIGH": "MEDIUM",
        "MEDIUM": "LOW",
        "LOW": "SLIGHT",
    },
    "MEDIUM": {
        "VERY_HIGH": "MAJOR",
        "HIGH": "HIGH",
        "RELATIVELY_HIGH": "MEDIUM",
        "MEDIUM": "LOW",
        "LOW": "SLIGHT",
    },
    "LOW": {
        "VERY_HIGH": "MEDIUM",
        "HIGH": "MEDIUM",
        "RELATIVELY_HIGH": "LOW",
        "MEDIUM": "SLIGHT",
        "LOW": "SLIGHT",
    },
}


ELECTRIC_HARM_RULE_CONFIG = {
    "version": 1,
    "industry": "ELECTRIC_POWER",
    "object_priority": ["NATIONAL_SECURITY", "PUBLIC_INTEREST", "LEGAL_RIGHTS"],
    "system_type_aliases": {
        "POWER_MONITORING_SYSTEM": "POWER_MONITORING_SYSTEM",
        "电力监控系统": "POWER_MONITORING_SYSTEM",
        "MANAGEMENT_INFO_SYSTEM": "MANAGEMENT_INFO_SYSTEM",
        "MANAGEMENT_SYSTEM": "MANAGEMENT_INFO_SYSTEM",
        "管理信息系统": "MANAGEMENT_INFO_SYSTEM",
        "COMMUNICATION_NETWORK_FACILITY": "COMMUNICATION_NETWORK_FACILITY",
        "通信网络设施": "COMMUNICATION_NETWORK_FACILITY",
        "DATA_RESOURCE": "DATA_RESOURCE",
        "数据资源": "DATA_RESOURCE",
        "OTHER_SYSTEM_PLATFORM": "OTHER_SYSTEM_PLATFORM",
        "OTHER_SYSTEM": "OTHER_SYSTEM_PLATFORM",
        "其他系统平台": "OTHER_SYSTEM_PLATFORM",
    },
    "object_names": {
        "NATIONAL_SECURITY": "国家安全",
        "PUBLIC_INTEREST": "社会秩序和公共利益",
        "LEGAL_RIGHTS": "其他公民、法人和组织的合法权益",
    },
    "damage_degree_names": {
        "GENERAL": "一般损害",
        "SERIOUS": "严重损害",
        "EXTREMELY_SERIOUS": "特别严重损害",
    },
    "system_categories": {
        "POWER_MONITORING_SYSTEM": {
            "name": "电力监控系统",
            "impact_degrees": {
                "LEGAL_RIGHTS": ["GENERAL", "SERIOUS", "EXTREMELY_SERIOUS"],
                "PUBLIC_INTEREST": ["GENERAL", "SERIOUS", "EXTREMELY_SERIOUS"],
                "NATIONAL_SECURITY": ["GENERAL", "SERIOUS"],
            },
        },
        "MANAGEMENT_INFO_SYSTEM": {
            "name": "管理信息系统",
            "impact_degrees": {
                "LEGAL_RIGHTS": ["GENERAL", "SERIOUS", "EXTREMELY_SERIOUS"],
                "PUBLIC_INTEREST": ["GENERAL", "SERIOUS", "EXTREMELY_SERIOUS"],
                "NATIONAL_SECURITY": ["GENERAL"],
            },
        },
        "COMMUNICATION_NETWORK_FACILITY": {
            "name": "通信网络设施",
            "impact_degrees": {
                "LEGAL_RIGHTS": ["GENERAL", "SERIOUS", "EXTREMELY_SERIOUS"],
                "PUBLIC_INTEREST": ["GENERAL", "SERIOUS"],
                "NATIONAL_SECURITY": ["GENERAL"],
            },
        },
        "DATA_RESOURCE": {
            "name": "数据资源",
            "impact_degrees": {
                "LEGAL_RIGHTS": ["GENERAL", "SERIOUS", "EXTREMELY_SERIOUS"],
                "PUBLIC_INTEREST": ["GENERAL", "SERIOUS", "EXTREMELY_SERIOUS"],
                "NATIONAL_SECURITY": ["GENERAL", "SERIOUS"],
            },
        },
        "OTHER_SYSTEM_PLATFORM": {
            "name": "其他系统平台",
            "impact_degrees": {
                "LEGAL_RIGHTS": ["GENERAL", "SERIOUS", "EXTREMELY_SERIOUS"],
                "PUBLIC_INTEREST": ["GENERAL", "SERIOUS"],
                "NATIONAL_SECURITY": [],
            },
        },
    },
    "protection_level_matrix": {
        "LEGAL_RIGHTS": {"GENERAL": 1, "SERIOUS": 2, "EXTREMELY_SERIOUS": 2},
        "PUBLIC_INTEREST": {"GENERAL": 2, "SERIOUS": 3, "EXTREMELY_SERIOUS": 4},
        "NATIONAL_SECURITY": {"GENERAL": 3, "SERIOUS": 4, "EXTREMELY_SERIOUS": 5},
    },
    "harm_level_by_protection_level": {
        "1": "LOW",
        "2": "MEDIUM",
        "3": "RELATIVELY_HIGH",
        "4": "HIGH",
        "5": "VERY_HIGH",
    },
    "damage_examples": {
        "NATIONAL_SECURITY": {
            "GENERAL": "使电网瓦解、发电机组停运、电力生产与供应中断，影响波及一个或多个地市的部分地区，明显影响社会安定。",
            "SERIOUS": "使电网瓦解、发电机组停运、电力生产与供应中断，影响波及一个或多个地市的大部分地区，对社会安定造成严重影响，明显影响国家安全。",
            "EXTREMELY_SERIOUS": "造成电网瓦解、发电机组停运，电力生产与供应中断，影响波及一个或多个省市的大部分地区，引起社会动荡，严重威胁国家安全。",
        },
        "PUBLIC_INTEREST": {
            "GENERAL": "使电力生产及供应面临明显的中断威胁，影响波及一个地市的部分地区，对公众利益造成一定危害，可能扰乱社会秩序。",
            "SERIOUS": "使电力生产及供应面临严重的中断威胁，影响波及一个或多个地市的部分地区，对公众利益造成严重危害，对社会秩序造成一定影响。",
            "EXTREMELY_SERIOUS": "使电网瓦解、发电机组停运、用电服务中断，影响波及一个或多个地市的大部分地区，严重扰乱社会秩序，对电力行业造成巨大经济损失，对公众利益造成特别严重危害。",
        },
        "LEGAL_RIGHTS": {
            "GENERAL": "对电力企业造成一定的经济损失，或对个别公民、法人或其它组织的利益造成较低的损害。",
            "SERIOUS": "对电力企业造成严重的经济损失，或对大量公民、法人或其它组织的利益造成特别严重的损害。",
            "EXTREMELY_SERIOUS": "对电力企业造成重大的经济损失，或对大量公民、法人或其它组织的利益造成特别严重的损害。",
        },
    },
}


ELECTRIC_HARM_LEVEL_RULES = [
    {
        "level": "VERY_HIGH",
        "description": "一旦发生数据安全风险，对国家安全、经济运行造成严重危害或特别严重危害，对社会稳定、公共利益造成特别严重危害。",
        "impact_object": "国家安全",
        "example": ELECTRIC_HARM_RULE_CONFIG["damage_examples"]["NATIONAL_SECURITY"]["EXTREMELY_SERIOUS"],
        "sort_order": 1,
    },
    {
        "level": "HIGH",
        "description": "一旦发生数据安全风险，对国家安全和经济运行产生危害，对社会秩序和公共利益产生严重危害。",
        "impact_object": "国家安全、社会秩序和公共利益",
        "example": ELECTRIC_HARM_RULE_CONFIG["damage_examples"]["NATIONAL_SECURITY"]["SERIOUS"],
        "sort_order": 2,
    },
    {
        "level": "RELATIVELY_HIGH",
        "description": "一旦发生数据安全风险，对国家安全和经济运行产生有限危害，对社会秩序和公共利益产生危害。",
        "impact_object": "国家安全、社会秩序和公共利益",
        "example": ELECTRIC_HARM_RULE_CONFIG["damage_examples"]["PUBLIC_INTEREST"]["SERIOUS"],
        "sort_order": 3,
    },
    {
        "level": "MEDIUM",
        "description": "一旦发生数据安全风险，对国家安全和经济运行不产生危害，对社会秩序和公共利益产生一般危害，对组织权益、组织自身运营产生危害。",
        "impact_object": "社会秩序和公共利益、其他公民、法人和组织的合法权益",
        "example": ELECTRIC_HARM_RULE_CONFIG["damage_examples"]["PUBLIC_INTEREST"]["GENERAL"],
        "sort_order": 4,
    },
    {
        "level": "LOW",
        "description": "一旦发生数据安全风险，对国家安全和经济运行、社会秩序和公共利益几乎不产生危害，对组织权益、组织自身运营、个人权益造成一般危害。",
        "impact_object": "其他公民、法人和组织的合法权益",
        "example": ELECTRIC_HARM_RULE_CONFIG["damage_examples"]["LEGAL_RIGHTS"]["GENERAL"],
        "sort_order": 5,
    },
]


def seed_default_data(excel_path: str, risk_source_template_path: str | None = None) -> dict:
    session = SessionLocal()
    created = {"templates": 0, "items": 0, "models": 0, "riskSourceTemplates": 0}

    template = session.get(AssessmentTemplate, "tpl-gb")
    if not template:
        template = AssessmentTemplate(
            id="tpl-gb",
            template_name="国标测评模板",
            template_type="NATIONAL",
            version=1,
            status="ENABLED",
            item_count=0,
        )
        session.add(template)
        created["templates"] += 1
    created["items"] = _seed_template_items(session, template, excel_path)
    template.item_count = created["items"]

    if not session.get(ScoreModel, "score-v1"):
        score_model = ScoreModel(
            id="score-v1",
            model_name="V1评分模型",
            model_type="PRESET",
            version=1,
            status="ENABLED",
            result_scores={
                "COMPLIANT": 1,
                "PARTIAL": 0.5,
                "NON_COMPLIANT": 0,
                "NOT_APPLICABLE": None,
            },
        )
        session.add(score_model)
        for level, min_score, max_score, include_min, include_max in DEFAULT_SCORE_RANGES:
            session.add(
                ScoreModelRange(
                    id=f"score-v1-{level.lower()}",
                    score_model_id="score-v1",
                    level=level,
                    min_score=min_score,
                    max_score=max_score,
                    include_min=include_min,
                    include_max=include_max,
                )
            )
        created["models"] += 1

    if not session.get(HarmModel, "harm-default"):
        session.add(
            HarmModel(
                id="harm-default",
                model_name="默认分析模型",
                version=1,
                status="ENABLED",
                description="Default harm model for phase-one integration.",
                rule_config=ELECTRIC_HARM_RULE_CONFIG,
            )
        )
        created["models"] += 1

    if not session.get(HarmModel, "harm-electric"):
        session.add(
            HarmModel(
                id="harm-electric",
                model_name="电力行业分析模型",
                version=1,
                status="ENABLED",
                description="Power industry harm model for knowledge page prototype tabs.",
                rule_config=ELECTRIC_HARM_RULE_CONFIG,
            )
        )
        created["models"] += 1

    matrix = session.get(RiskMatrix, "matrix-v1")
    if not matrix:
        session.add(
            RiskMatrix(
                id="matrix-v1",
                matrix_name="V1评价矩阵",
                version=1,
                status="ENABLED",
                remark="默认 3x5 风险评价矩阵",
                matrix_json=DEFAULT_RISK_MATRIX,
            )
        )
        created["models"] += 1
    elif not _risk_matrix_has_required_axes(matrix.matrix_json):
        matrix.matrix_json = DEFAULT_RISK_MATRIX
        matrix.remark = matrix.remark or "默认 3x5 风险评价矩阵"

    _seed_auxiliary_knowledge(session)
    created["riskSourceTemplates"] = _seed_risk_source_templates(session, risk_source_template_path or _default_risk_source_template_path())
    _backfill_existing_assessment_item_ids(session)
    _backfill_existing_risk_template_fields(session)
    session.commit()
    return created


def _seed_auxiliary_knowledge(session) -> None:
    if not session.get(RemediationSuggestionTemplate, "remtpl-default"):
        session.add(
            RemediationSuggestionTemplate(
                id="remtpl-default",
                suggestion_title="完善数据安全管理制度",
                risk_level="MEDIUM",
                risk_type="DATA_LEAKAGE",
                suggestion_content="建议补充数据分类分级、访问控制、日志审计和应急响应等管理制度并定期复核执行情况。",
                status="ENABLED",
            )
        )

    _seed_harm_model_knowledge(session, "harm-default", "harmrule-default")
    _seed_harm_model_knowledge(session, "harm-electric", "harmrule-electric")


def _risk_matrix_has_required_axes(matrix_json: dict | None) -> bool:
    """检查风险评价矩阵是否至少包含 3 行发生可能性和 5 列危害程度。"""
    if not isinstance(matrix_json, dict):
        return False
    required_rows = set(DEFAULT_RISK_MATRIX.keys())
    required_columns = set(next(iter(DEFAULT_RISK_MATRIX.values())).keys())
    if set(matrix_json.keys()) != required_rows:
        return False
    for possibility_level in required_rows:
        row = matrix_json.get(possibility_level)
        if not isinstance(row, dict) or set(row.keys()) != required_columns:
            return False
    return True


def _seed_harm_model_knowledge(session, model_id: str, rule_id_prefix: str) -> None:
    model = session.get(HarmModel, model_id)
    if not model:
        return
    if not model.rule_config:
        model.rule_config = ELECTRIC_HARM_RULE_CONFIG
    for rule in ELECTRIC_HARM_LEVEL_RULES:
        rule_id = f"{rule_id_prefix}-{rule['level'].lower().replace('_', '-')}"
        existing = session.get(HarmModelRule, rule_id)
        if not existing:
            existing = HarmModelRule(id=rule_id, harm_model_id=model_id)
            session.add(existing)
        existing.level = rule["level"]
        existing.description = rule["description"]
        existing.impact_object = rule["impact_object"]
        existing.example = rule["example"]
        existing.judgement_steps = ["明确被评估对象类别", "确定受侵害客体和侵害程度", "匹配数据安全保护等级", "匹配风险危害程度等级"]
        existing.sort_order = rule["sort_order"]

def _seed_template_items(session, template: AssessmentTemplate, excel_path: str) -> int:
    path = Path(excel_path)
    if not path.exists():
        return template.item_count or 0

    workbook = load_workbook(path, read_only=True, data_only=True)
    sort_order = 0
    active_ids = set()
    existing_items = session.query(AssessmentTemplateItem).filter(AssessmentTemplateItem.template_id == template.id).all()
    existing_by_id = {item.id: item for item in existing_items}
    existing_by_standard_id = {
        item.standard_item_id: item
        for item in existing_items
        if item.standard_item_id
    }
    existing_by_assessment_item_id = {
        item.assessment_item_id: item
        for item in existing_items
        if item.assessment_item_id
    }
    existing_by_context = {
        _assessment_context_key(item.sheet_name, item.category, item.subcategory, item.check_point): item
        for item in existing_items
    }
    for sheet in workbook.worksheets:
        sheet_name = _normalize_assessment_sheet_name(sheet.title)
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        header_map = _assessment_header_map(header_row or ())
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or not any(value not in (None, "") for value in row):
                continue
            check_point = _assessment_cell(row, header_map, "评估项", 3)
            if not check_point:
                continue
            sort_order += 1
            category = _assessment_cell(row, header_map, "评估类别", 1)
            subcategory = _assessment_cell(row, header_map, "评估子类", 2)
            assessment_item_id = _assessment_cell(row, header_map, "评估项ID", None)
            standard_item_id = _assessment_cell(row, header_map, "标准项ID", 9)
            item_code = _assessment_cell(row, header_map, "整体序号", 0) or str(sort_order)
            item_id = _assessment_template_item_id(sheet_name, category, subcategory, check_point, standard_item_id, assessment_item_id, sort_order)
            item = (
                existing_by_id.get(item_id)
                or (existing_by_standard_id.get(standard_item_id) if standard_item_id else None)
                or (existing_by_assessment_item_id.get(assessment_item_id) if assessment_item_id else None)
                or existing_by_context.get(
                    _assessment_context_key(sheet_name, category, subcategory, check_point)
                )
            )
            if not item:
                item = AssessmentTemplateItem(id=item_id, template_id=template.id)
                session.add(item)
                existing_items.append(item)
                existing_by_id[item.id] = item
            item.deleted = False
            item.sheet_name = sheet_name
            item.category = category
            item.subcategory = subcategory
            item.category_id = _category_id(sheet_name, category, subcategory)
            item.item_code = item_code
            item.assessment_item_id = assessment_item_id
            item.check_point = check_point
            item.standard_item_id = standard_item_id
            item.sort_order = sort_order
            if item.standard_item_id:
                existing_by_standard_id[item.standard_item_id] = item
            if item.assessment_item_id:
                existing_by_assessment_item_id[item.assessment_item_id] = item
            existing_by_context[_assessment_context_key(item.sheet_name, item.category, item.subcategory, item.check_point)] = item
            active_ids.add(item.id)

    if active_ids:
        for item in existing_items:
            if item.id not in active_ids:
                item.deleted = True
    return len(active_ids) or template.item_count or 0


def _assessment_header_map(header_row: tuple) -> dict[str, int]:
    return {str(value or "").strip(): index for index, value in enumerate(header_row) if str(value or "").strip()}


def _assessment_cell(row: tuple, header_map: dict[str, int], header: str, fallback_index: int | None) -> str:
    index = header_map.get(header)
    if index is None:
        index = fallback_index
    if index is None or index >= len(row):
        return ""
    return _cell_to_text(row[index])


def _cell_to_text(value) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalize_assessment_sheet_name(sheet_name: str) -> str:
    return re.sub(r"\d+(?:[-+]\d+)*$", "", str(sheet_name or "")).strip() or str(sheet_name or "").strip()


def _assessment_template_item_id(
    sheet_name: str,
    category: str,
    subcategory: str,
    check_point: str,
    standard_item_id: str,
    assessment_item_id: str,
    sort_order: int,
) -> str:
    if standard_item_id:
        return f"tplitem-{standard_item_id}"
    if assessment_item_id:
        return f"tplitem-{assessment_item_id.lower()}"
    digest = hashlib.sha1("|".join([sheet_name, category, subcategory, check_point]).encode("utf-8")).hexdigest()
    return f"tplitem-{sort_order}-{digest[:10]}"


def _assessment_context_key(sheet_name: str | None, category: str | None, subcategory: str | None, check_point: str | None) -> tuple[str, str, str, str]:
    return (
        _normalize_assessment_key(sheet_name),
        _normalize_assessment_key(category),
        _normalize_assessment_key(subcategory),
        _normalize_assessment_key(check_point),
    )


def _normalize_assessment_key(value: str | None) -> str:
    return " ".join(str(value or "").split())


def _find_existing_template_item(
    session,
    template_id: str,
    sheet_name: str,
    category: str,
    subcategory: str,
    check_point: str,
    standard_item_id: str,
    assessment_item_id: str,
) -> AssessmentTemplateItem | None:
    query = session.query(AssessmentTemplateItem).filter(AssessmentTemplateItem.template_id == template_id)
    if standard_item_id:
        item = query.filter(AssessmentTemplateItem.standard_item_id == standard_item_id).first()
        if item:
            return item
    if assessment_item_id:
        item = query.filter(AssessmentTemplateItem.assessment_item_id == assessment_item_id).first()
        if item:
            return item
    return query.filter(
        AssessmentTemplateItem.sheet_name == sheet_name,
        AssessmentTemplateItem.category == category,
        AssessmentTemplateItem.subcategory == subcategory,
        AssessmentTemplateItem.check_point == check_point,
    ).first()


def _backfill_existing_assessment_item_ids(session) -> None:
    session.flush()
    session.execute(
        text(
            """
            UPDATE project_assessment_item
            SET assessment_item_id = (
                SELECT assessment_template_item.assessment_item_id
                FROM assessment_template_item
                WHERE assessment_template_item.id = project_assessment_item.template_item_id
            )
            WHERE deleted = false
              AND template_item_id IS NOT NULL
              AND (
                assessment_item_id IS NULL
                OR assessment_item_id = ''
                OR assessment_item_id <> (
                    SELECT assessment_template_item.assessment_item_id
                    FROM assessment_template_item
                    WHERE assessment_template_item.id = project_assessment_item.template_item_id
                )
              )
              AND (
                SELECT assessment_template_item.assessment_item_id
                FROM assessment_template_item
                WHERE assessment_template_item.id = project_assessment_item.template_item_id
              ) IS NOT NULL
              AND (
                SELECT assessment_template_item.assessment_item_id
                FROM assessment_template_item
                WHERE assessment_template_item.id = project_assessment_item.template_item_id
              ) <> ''
            """
        )
    )
    session.execute(
        text(
            """
            UPDATE project_risk_summary_record
            SET assessment_item_id = (
                SELECT project_assessment_item.assessment_item_id
                FROM project_assessment_item
                WHERE project_assessment_item.id = project_risk_summary_record.evaluation_item_id
            )
            WHERE deleted = false
              AND evaluation_item_id IS NOT NULL
              AND (
                assessment_item_id IS NULL
                OR assessment_item_id = ''
                OR assessment_item_id <> (
                    SELECT project_assessment_item.assessment_item_id
                    FROM project_assessment_item
                    WHERE project_assessment_item.id = project_risk_summary_record.evaluation_item_id
                )
              )
              AND (
                SELECT project_assessment_item.assessment_item_id
                FROM project_assessment_item
                WHERE project_assessment_item.id = project_risk_summary_record.evaluation_item_id
              ) IS NOT NULL
              AND (
                SELECT project_assessment_item.assessment_item_id
                FROM project_assessment_item
                WHERE project_assessment_item.id = project_risk_summary_record.evaluation_item_id
              ) <> ''
            """
        )
    )


def _backfill_existing_risk_template_fields(session) -> None:
    templates = (
        session.query(RiskSourceTemplate)
        .filter(RiskSourceTemplate.deleted.is_(False))
        .order_by(RiskSourceTemplate.sort_order.asc())
        .all()
    )
    templates_by_key = {}
    for template in templates:
        key = _risk_template_key(template.sheet_name, template.category, template.subcategory, template.assessment_item)
        if key[3] and key not in templates_by_key:
            templates_by_key[key] = template

    rows = (
        session.query(ProjectRiskSummaryRecord, ProjectAssessmentItem)
        .join(ProjectAssessmentItem, ProjectAssessmentItem.id == ProjectRiskSummaryRecord.evaluation_item_id)
        .filter(ProjectRiskSummaryRecord.deleted.is_(False), ProjectAssessmentItem.deleted.is_(False))
        .all()
    )
    for record, item in rows:
        key = _risk_template_key(
            item.sheet_name,
            record.assessment_category or item.category,
            record.assessment_subcategory or item.subcategory,
            record.check_point or item.check_point,
        )
        template = templates_by_key.get(key)
        if not template:
            continue
        record.risk_source_type = template.risk_source_type


def _seed_risk_source_templates(session, excel_path: str) -> int:
    path = Path(excel_path)
    if not path.exists():
        return 0

    workbook = load_workbook(path, read_only=True, data_only=True)
    saved = 0
    sort_order = 0
    for sheet in workbook.worksheets:
        headers = [str(value or "").strip() for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
        header_map = {header: index for index, header in enumerate(headers)}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or not any(value not in (None, "") for value in row):
                continue
            sort_order += 1
            category = _risk_template_cell(row, header_map, "评估类别")
            subcategory = _risk_template_cell(row, header_map, "评估子类")
            assessment_item = _risk_template_cell(row, header_map, "评估项")
            if not category or not subcategory or not assessment_item:
                continue
            template = _get_risk_source_template(session, sheet.title, category, subcategory, assessment_item)
            if not template:
                template = RiskSourceTemplate(
                    id=_risk_source_template_id(sheet.title, category, subcategory, assessment_item),
                    sheet_name=sheet.title,
                    category=category,
                    subcategory=subcategory,
                    assessment_item=assessment_item,
                )
                session.add(template)
                saved += 1
            template.evaluation_record = _risk_template_cell(row, header_map, "评估记录")
            template.evaluation_result = _risk_template_cell(row, header_map, "评估结果")
            template.risk_description = _risk_template_cell(row, header_map, "问题描述")
            template.remediation_suggestion = _risk_template_cell(row, header_map, "整改建议")
            template.risk_source_description = _risk_template_cell(row, header_map, "常见风险源")
            template.risk_source_type = _risk_template_cell(row, header_map, "风险源类型")
            template.risk_types = _split_risk_types(_risk_template_cell(row, header_map, "风险类型"))
            template.sort_order = sort_order
    return saved


def _category_id(sheet_name: str, category: str, subcategory: str) -> str:
    return "|".join([sheet_name or "-", category or "-", subcategory or "-"])


def _default_risk_source_template_path() -> str:
    return str(Path(__file__).resolve().parents[2] / "doc" / "国标风险源模版.xlsx")


def _risk_template_cell(row: tuple, header_map: dict[str, int], header: str) -> str:
    index = header_map.get(header)
    if index is None or index >= len(row):
        return ""
    value = row[index]
    return str(value or "").strip()


def _get_risk_source_template(
    session,
    sheet_name: str,
    category: str,
    subcategory: str,
    assessment_item: str,
) -> RiskSourceTemplate | None:
    return (
        session.query(RiskSourceTemplate)
        .filter(
            RiskSourceTemplate.sheet_name == sheet_name,
            RiskSourceTemplate.category == category,
            RiskSourceTemplate.subcategory == subcategory,
            RiskSourceTemplate.assessment_item == assessment_item,
            RiskSourceTemplate.deleted.is_(False),
        )
        .first()
    )


def _risk_source_template_id(sheet_name: str, category: str, subcategory: str, assessment_item: str) -> str:
    digest = hashlib.sha1("|".join([sheet_name, category, subcategory, assessment_item]).encode("utf-8")).hexdigest()
    return f"rstpl-{digest[:16]}"


def _risk_template_key(sheet_name: str | None, category: str | None, subcategory: str | None, assessment_item: str | None) -> tuple[str, str, str, str]:
    return (
        _normalize_risk_template_key(sheet_name),
        _normalize_risk_template_key(category),
        _normalize_risk_template_key(subcategory),
        _normalize_risk_template_key(assessment_item),
    )


def _normalize_risk_template_key(value: str | None) -> str:
    return " ".join(str(value or "").split())


def _split_risk_types(value: str) -> list[str]:
    if not value:
        return []
    return [item.strip() for item in value.replace("，", ",").split(",") if item.strip()]
